"""
Telegram Bot Quản Lý Chi Tiêu - ExpenseBot Super Multimodal AI Edition
Sử dụng Groq AI (Llama 3.3 70B, Vision, Whisper) + Smart Pattern Matching Fallback
Phiên bản Super Multimodal với Text, Voice, Vision, RAG Context

Copyright (c) 2025 Lộc
All rights reserved.

This software is proprietary and confidential. Unauthorized copying, modification,
distribution, or use of this software, via any medium, is strictly prohibited.
"""

import os
import re
import json
import logging
import io
import base64
import tempfile
import requests
from datetime import datetime, timedelta, time as dt_time
from collections import defaultdict
from dotenv import load_dotenv
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, CommandHandler, MessageHandler, CallbackQueryHandler, filters, ContextTypes, JobQueue
from telegram.constants import ParseMode
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import pandas as pd
import matplotlib
matplotlib.use('Agg')  # Thread-safe backend
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from keep_alive import keep_alive
from services import (
    init_google_sheets,
    save_expenses_to_sheet,
    calculate_weekly_spend,
    get_financial_context,
    get_expense_report,
    get_worksheet,
    google_search,
    generate_image,
    generate_vietqr_url,
    classify_intent_with_ai,
    find_expense_by_name,
    delete_expense_by_row_index
)

# Load biến môi trường từ file .env
load_dotenv()

# ==================== CẤU HÌNH LOGGING ====================
logging.basicConfig(
    format='%(asctime)s - [%(levelname)s] - %(message)s',
    level=logging.INFO,
    handlers=[logging.StreamHandler()]
)
logger = logging.getLogger(__name__)

# Import Text-to-Speech
try:
    from gtts import gTTS
    from pydub import AudioSegment
    TTS_AVAILABLE = True
except ImportError:
    TTS_AVAILABLE = False
    logger.warning("⚠️ gTTS hoặc pydub chưa được cài đặt. Voice Reply sẽ bị tắt.")

# Import Groq AI
try:
    from groq import Groq
    GROQ_AVAILABLE = True
except ImportError:
    GROQ_AVAILABLE = False
    logger.warning("⚠️ Groq library not installed. AI features will be disabled.")

# ==================== CẤU HÌNH ====================
TELEGRAM_TOKEN = os.getenv('BOT_TOKEN', '')
CREDENTIALS_FILE = 'credentials.json'

# Google Search API Configuration (để kiểm tra trong error handling)
GOOGLE_SEARCH_API_KEY = os.getenv('GOOGLE_SEARCH_API_KEY', '')
GOOGLE_CSE_ID = os.getenv('GOOGLE_CSE_ID', '')
SHEET_NAME = 'QuanLyChiTieu'
SHEET_ID = os.getenv('GOOGLE_SHEET_ID', '')

# ==================== CẤU HÌNH GROQ AI ====================
GROQ_API_KEY = os.getenv('GROQ_API_KEY', '')
groq_client = None
groq_disabled = False  # Flag để tạm thời disable Groq nếu quota hết
GROQ_PRIORITY = True  # Ưu tiên sử dụng Groq AI

if GROQ_AVAILABLE and GROQ_API_KEY:
    try:
        groq_client = Groq(api_key=GROQ_API_KEY)
        logger.info("✅ Đã khởi tạo Groq AI client (Llama 3)")
    except Exception as e:
        logger.warning(f"⚠️ Không thể khởi tạo Groq client: {e}")
        groq_client = None
elif not GROQ_AVAILABLE:
    logger.warning("⚠️ Groq library chưa được cài đặt. Chạy: pip install groq")
elif not GROQ_API_KEY:
    logger.info("ℹ️ GROQ_API_KEY chưa được cấu hình. Bot sẽ sử dụng Regex fallback.")
else:
    groq_client = None

# ==================== CẤU HÌNH NGÂN SÁCH TUẦN ====================
WEEKLY_LIMIT = 700000  # 700 nghìn đồng/tuần

# ==================== LƯU TRỮ REMINDER ====================
REMINDER_FILE = 'reminders.json'
user_reminders = {}  # {user_id: {'hour': int, 'minute': int}}

# Load reminders từ file nếu có
def load_reminders():
    """Load reminders từ file JSON"""
    global user_reminders
    try:
        if os.path.exists(REMINDER_FILE):
            with open(REMINDER_FILE, 'r', encoding='utf-8') as f:
                user_reminders = json.load(f)
                logger.info(f"✅ Đã load {len(user_reminders)} reminders từ file")
    except Exception as e:
        logger.warning(f"⚠️ Không thể load reminders: {e}")
        user_reminders = {}

def save_reminders():
    """Lưu reminders vào file JSON"""
    try:
        with open(REMINDER_FILE, 'w', encoding='utf-8') as f:
            json.dump(user_reminders, f, ensure_ascii=False, indent=2)
        logger.info("✅ Đã lưu reminders vào file")
    except Exception as e:
        logger.error(f"❌ Không thể lưu reminders: {e}")

# Load reminders khi khởi động
load_reminders()

# ==================== CHAT MEMORY (SHORT-TERM MEMORY) ====================
chat_memory = {}  # {user_id: [{'role': 'user'|'bot', 'content': '...'}]}
MAX_MEMORY_MESSAGES = 6  # Tối đa 6 tin nhắn (3 user, 3 bot)

def format_chat_history(user_id: int) -> str:
    """
    Format lịch sử chat của user thành chuỗi text
    Trả về chuỗi rỗng nếu không có lịch sử
    """
    if user_id not in chat_memory or not chat_memory[user_id]:
        return ""
    
    history_lines = ["Lịch sử trò chuyện:"]
    for msg in chat_memory[user_id]:
        role = msg.get('role', '')
        content = msg.get('content', '')
        if role == 'user':
            history_lines.append(f"User: {content}")
        elif role == 'bot':
            history_lines.append(f"Bot: {content}")
    
    return "\n".join(history_lines)

def add_to_memory(user_id: int, role: str, content: str):
    """
    Thêm tin nhắn vào memory của user
    Tự động giới hạn tối đa MAX_MEMORY_MESSAGES
    """
    if user_id not in chat_memory:
        chat_memory[user_id] = []
    
    # Thêm tin nhắn mới
    chat_memory[user_id].append({
        'role': role,
        'content': content
    })
    
    # Giới hạn số lượng tin nhắn
    if len(chat_memory[user_id]) > MAX_MEMORY_MESSAGES:
        chat_memory[user_id] = chat_memory[user_id][-MAX_MEMORY_MESSAGES:]
    
    logger.info(f"💾 Đã lưu vào memory: {role} - {content[:50]}... (Total: {len(chat_memory[user_id])} messages)")

logger.info("=" * 60)
logger.info("KHỞI ĐỘNG BOT QUẢN LÝ CHI TIÊU (Enterprise Edition)")
logger.info("=" * 60)
logger.info(f"💰 Hạn mức tuần: {WEEKLY_LIMIT:,}đ")
logger.info("📊 Tính năng Enterprise: Biểu đồ, Xuất Excel, Reminder, Bill Splitter")

if not TELEGRAM_TOKEN:
    logger.critical("❌ CRITICAL ERROR: TELEGRAM_TOKEN không được tìm thấy!")
    raise ValueError("TELEGRAM_TOKEN không được tìm thấy!")
else:
    logger.info("✅ TELEGRAM_TOKEN: Đã tìm thấy")

# ==================== TỪ KHÓA LÃNG PHÍ (Cảnh Sát Chi Tiêu) ====================
WASTEFUL_KEYWORDS = [
    'game', 'nạp', 'nap', 'skin', 'gacha', 'trà sữa', 'tra sua', 'toco', 'mixue', 
    'phim', 'netflix', 'đồ chơi', 'do choi', 'mô hình', 'mo hinh', 'nhậu', 'nhau',
    'pubg', 'lol', 'liên quân', 'lien quan', 'mobile legend', 'genshin', 'top up',
    'thẻ game', 'the game', 'card', 'gift code', 'code', 'vip', 'premium'
]

WASTEFUL_WARNINGS = [
    "Tiền không phải lá mít đâu nhé! 💸",
    "Lại tốn tiền vào cái này rồi, chán thanh niên! 😒",
    "Bớt bớt lại đi, cuối tháng ăn mì gói bây giờ! 🍜",
    "Tiêu tiền như nước, rồi lại than nghèo! 💧",
    "Cẩn thận kẻo hết tiền trước khi hết tháng! ⚠️",
    "Nhớ tiết kiệm một chút, đừng phung phí quá! 💰",
    "Lại chi tiêu không cần thiết rồi, cẩn thận nhé! 🚨",
    "Tiền kiếm được khó lắm, đừng vứt đi như vậy! 😤",
    "Có tiền thì tiêu, không có tiền thì... than! 😅",
    "Nhớ mục tiêu tiết kiệm của mình nhé! 🎯"
]

logger.info("✅ Đã tải từ khóa lãng phí và cảnh báo")

# ==================== TỪ ĐIỂN PHÂN LOẠI TỰ ĐỘNG ====================
CATEGORY_KEYWORDS = {
    'Ăn uống': [
        'phở', 'pho', 'cơm', 'com', 'bún', 'bun', 'nước', 'nuoc', 'cf', 'cafe', 'cà phê', 'ca phe',
        'trà', 'tra', 'chè', 'che', 'bánh', 'banh', 'mì', 'mi', 'bánh mì', 'banh mi', 'xôi', 'xoi',
        'cháo', 'chao', 'súp', 'sup', 'lẩu', 'lau', 'nướng', 'nuong', 'gà', 'ga', 'thịt', 'thit',
        'cá', 'ca', 'tôm', 'tom', 'rau', 'đồ ăn', 'do an', 'ăn', 'an', 'uống', 'uong', 'nước uống',
        'nuoc uong', 'sữa', 'sua', 'kem', 'bánh kẹo', 'banh keo', 'snack', 'kẹo', 'keo'
    ],
    'Di chuyển': [
        'xăng', 'xang', 'xe', 'grab', 'be', 'uber', 'taxi', 'gửi xe', 'gui xe', 'đỗ xe', 'do xe',
        'bãi xe', 'bai xe', 'vé', 've', 'ticket', 'máy bay', 'may bay', 'tàu', 'tau', 'xe bus',
        'xe buýt', 'xe buyt', 'đi lại', 'di lai', 'vận chuyển', 'van chuyen', 'ship', 'giao hàng',
        'giao hang', 'đi', 'di', 'về', 've', 'đi về', 'di ve'
    ],
    'Học tập': [
        'vở', 'vo', 'sách', 'sach', 'bút', 'but', 'học', 'hoc', 'sách giáo khoa', 'sach giao khoa',
        'tài liệu', 'tai lieu', 'photocopy', 'photo', 'in', 'mực', 'muc', 'bút chì', 'but chi',
        'thước', 'thuoc', 'compa', 'máy tính', 'may tinh', 'calculator', 'học phí', 'hoc phi',
        'phí học', 'phi hoc', 'đăng ký', 'dang ky', 'đăng kí', 'dang ki', 'khóa học', 'khoa hoc'
    ]
}

logger.info("✅ Đã tải từ điển phân loại tự động")

# ==================== HÀM XỬ LÝ THÔNG MINH ====================
def parse_amount(text: str) -> tuple:
    """Tìm và chuyển đổi số tiền từ text. Trả về: (amount, positions)"""
    text_lower = text.lower()
    patterns = [
        (r'(\d+(?:\.\d+)?)\s*tr(?:iệu)?', 1000000),
        (r'(\d+(?:\.\d+)?)\s*k(?:ilo)?', 1000),
        (r'(\d+(?:\.\d+)?)\s*ng(?:àn)?', 1000),
        (r'(\d+(?:\.\d+)?)\s*nghìn', 1000),
        (r'(\d+(?:\.\d+)?)\s*000', 1),
        (r'(\d+(?:\.\d+)?)\s*d(?:ồng)?', 1),
        (r'(\d+(?:\.\d+)?)\s*đ', 1),
        (r'(\d{4,})', 1),
    ]
    
    amounts_found = []
    for pattern, multiplier in patterns:
        matches = re.finditer(pattern, text_lower)
        for match in matches:
            try:
                number = float(match.group(1))
                amount = int(number * multiplier)
                amounts_found.append((amount, match.start(), match.end()))
            except:
                continue
    
    if amounts_found:
        amounts_found.sort(key=lambda x: x[0], reverse=True)
        amount = amounts_found[0][0]
        logger.info(f"💰 Tìm thấy số tiền: {amount:,}đ")
        return amount, amounts_found
    
    return 0, []


def extract_item_name(text: str, amount_positions: list) -> str:
    """Trích xuất tên món từ text, loại bỏ phần số tiền"""
    text_cleaned = text
    for amount, start, end in sorted(amount_positions, key=lambda x: x[1], reverse=True):
        text_cleaned = text_cleaned[:start] + text_cleaned[end:]
    
    text_cleaned = text_cleaned.strip()
    remove_words = ['nay', 'hôm nay', 'hom nay', 'vừa', 'vua', 'mới', 'moi', 'làm', 'lam', 
                    'ăn', 'an', 'uống', 'uong', 'mua', 'chi', 'tiêu', 'tieu', 'ngon', 'quá', 'qua']
    
    words = text_cleaned.split()
    words_cleaned = [w for w in words if w.lower() not in remove_words]
    item_name = ' '.join(words_cleaned).strip()
    item_name = re.sub(r'\d+', '', item_name).strip()
    
    if len(item_name) < 2:
        if amount_positions:
            first_amount_start = min(pos[1] for pos in amount_positions)
            item_name = text[:first_amount_start].strip()
        else:
            item_name = text.strip()
    
    item_name = re.sub(r'[^\w\s]', ' ', item_name)
    item_name = ' '.join(item_name.split())
    
    if not item_name:
        item_name = "Không xác định"
    
    logger.info(f"📝 Tên món trích xuất: {item_name}")
    return item_name


def auto_categorize(item_name: str) -> str:
    """Tự động phân loại dựa trên từ khóa trong tên món"""
    item_lower = item_name.lower()
    item_normalized = item_lower
    
    for category, keywords in CATEGORY_KEYWORDS.items():
        for keyword in keywords:
            if keyword in item_normalized:
                logger.info(f"🏷️ Phân loại: {category} (từ khóa: {keyword})")
                return category
    
    logger.info(f"🏷️ Phân loại: Khác")
    return "Khác"


def has_amount(text: str) -> bool:
    """
    Kiểm tra xem text có chứa số tiền hay không
    Trả về True nếu tìm thấy số tiền, False nếu không
    """
    amount, _ = parse_amount(text)
    return amount > 0


def parse_single_item(text: str) -> dict:
    """Parse một món đơn lẻ"""
    amount, amount_positions = parse_amount(text)
    if amount == 0:
        raise ValueError("Không tìm thấy số tiền")
    
    item_name = extract_item_name(text, amount_positions)
    category = auto_categorize(item_name)
    
    return {
        'item': item_name,
        'amount': amount,
        'category': category
    }


def parse_with_groq(input_data, context_data: str = "", input_type: str = 'text', chat_history: str = "") -> dict:
    """
    Bộ Não Trung Tâm - Xử lý đa modal với Groq AI
    - input_data: Text hoặc nội dung ảnh (base64)
    - context_data: Financial context từ Google Sheet
    - input_type: 'text', 'image', 'voice'
    - chat_history: Lịch sử trò chuyện gần nhất
    Trả về dict với:
    - type: "expense" hoặc "chat"
    - expenses: list (nếu type == "expense")
    - message: str (nếu type == "expense")
    - response: str (nếu type == "chat")
    """
    if not groq_client:
        raise Exception("Groq client không khả dụng")
    
    logger.info("=" * 60)
    logger.info(f"🤖 ĐANG SỬ DỤNG GROQ AI - Input Type: {input_type}")
    logger.info("=" * 60)
    
    # Lấy thời gian hiện tại để xử lý backdated entry
    current_time = datetime.now()
    current_time_str = current_time.strftime('%Y-%m-%d %H:%M:%S')
    current_date_str = current_time.strftime('%d/%m/%Y')
    
    # Chọn model dựa trên input type
    if input_type == 'image':
        model = "llama-3.2-90b-vision-preview"
        logger.info("📷 Sử dụng Vision Model (90B)")
    else:
        model = "llama-3.3-70b-versatile"
        logger.info("💬 Sử dụng Text Model")
    
    # System Prompt - Tách biệt context khỏi system prompt
    if input_type == 'image':
        # System Prompt cho Vision (đọc hóa đơn) - Tối ưu cho model 90B
        system_prompt = (
            "Bạn là AI Vision chuyên đọc hóa đơn tiếng Việt.\n"
            "Nhiệm vụ: Trích xuất danh sách món ăn và giá tiền từ ảnh hóa đơn.\n\n"
            "Output JSON format:\n"
            "{\"type\": \"expense\", \"expenses\": [{\"item\": \"Tên món\", \"amount\": 123000, \"category\": \"Ăn uống\"}], \"message\": \"Đã đọc hóa đơn...\"}\n\n"
            "Quy tắc:\n"
            "- Nếu ảnh mờ hoặc không phải hóa đơn, trả về type: \"chat\" và nhắc user chụp lại.\n"
            "- Ưu tiên tìm dòng \"Tổng cộng\" (Total) nếu danh sách món quá dài hoặc mờ.\n"
            "- Chỉ lấy các món có giá tiền rõ ràng.\n"
            "- Bỏ qua ngày giờ, địa chỉ quán, thông tin không liên quan.\n"
            "- Category: Dựa vào tên món (phở/cơm/bún -> Ăn uống, xăng/grab -> Di chuyển, sách/vở -> Học tập, còn lại -> Khác).\n\n"
            "LUÔN trả về JSON chuẩn. Không markdown."
        )
    else:
        # System Prompt cho Text - Cải thiện phân loại + Chat History + Backdated Entry + Google Search
        system_prompt = (
            "Bạn là Trợ lý AI thông minh kiêm thư ký riêng của Lộc. "
            "Bạn trả lời ngắn gọn, chuyên nghiệp nhưng thân thiện. "
            "Luôn sẵn sàng giúp đỡ và hỗ trợ.\n"
            f"Thời gian hiện tại của hệ thống là: {current_time_str} (Ngày: {current_date_str}).\n"
            "Dữ liệu hệ thống và lịch sử trò chuyện sẽ được cung cấp trong user message.\n\n"
            "PHÂN LOẠI INPUT (QUAN TRỌNG):\n\n"
            "1. **TYPE: \"expense\"** (Chỉ khi User nhập khoản chi MỚI):\n"
            "   - VD: \"phở 50k\", \"đổ xăng 200k\", \"mua rau 20k\", \"ăn trưa 35k, cafe 25k\".\n"
            "   - Đặc điểm: Có tên món + số tiền, là hành động CHI TIÊU MỚI.\n"
            "   - Output: {\"type\": \"expense\", \"expenses\": [{\"item\": \"tên món\", \"amount\": số_tiền_int, \"category\": \"Ăn uống/Di chuyển/Học tập/Khác\", \"date\": \"DD/MM/YYYY\" hoặc null}], \"message\": \"...\", \"image_prompt\": \"...\" (tùy chọn)}\n"
            "   - **image_prompt** (Tùy chọn): Nếu user tiêu hoang (>500k hoặc game/trà sữa), hãy thêm field này với prompt mô tả cảnh nghèo khổ/hài hước bằng tiếng Anh (VD: \"poor student eating instant noodles, anime style\").\n\n"
            "2. **TYPE: \"search\"** (Khi User hỏi về dữ liệu thực tế cần tìm kiếm):\n"
            "   - VD: \"Giá vàng hôm nay\", \"Ai là tổng thống Mỹ\", \"Thời tiết Hà Nội\", \"Giá xăng hôm nay\", \"Tin tức mới nhất\".\n"
            "   - Đặc điểm: Câu hỏi cần dữ liệu thực tế, cập nhật, hoặc thông tin không có trong hệ thống.\n"
            "   - Output: {\"type\": \"search\", \"query\": \"từ khóa tìm kiếm ngắn gọn\"}\n"
            "   - Lưu ý: Đừng trả lời bừa. Nếu không chắc chắn, hãy yêu cầu tìm kiếm.\n\n"
            "3. **TYPE: \"qr_request\"** (Khi User yêu cầu tạo mã QR chuyển khoản):\n"
            "   - VD: \"tạo mã qr 50k\", \"tạo cho tôi cái mã qr mệnh giá 20k nội dung là tra no\", \"qr code 100k tiền cafe\", \"mã chuyển khoản 500k\".\n"
            "   - Đặc điểm: User muốn tạo mã QR để nhận tiền chuyển khoản, có số tiền và nội dung (tùy chọn).\n"
            "   - Output: {\"type\": \"qr_request\", \"amount\": số_tiền_int, \"content\": \"nội dung chuyển khoản\" hoặc \"\"}\n"
            "   - Lưu ý: Phải trích xuất số tiền từ text (xử lý 'k', 'tr', 'ng', 'nghìn', 'triệu'). Nếu không có nội dung, để content = \"\".\n\n"
            "4. **TYPE: \"chat\"** (Khi User hỏi về dữ liệu hệ thống, tra cứu, tâm sự, hoặc nói chuyện bình thường):\n"
            "   - VD: \"hôm nay tiêu bao nhiêu?\", \"tài chính thế nào?\", \"còn bao nhiêu tiền?\", \"danh sách chi tiêu hôm nay\", \"chào em\", \"cảm ơn\", \"front end là gì\", \"hướng dẫn tôi học\".\n"
            "   - Đặc điểm: Là câu HỎI về dữ liệu hệ thống, TRA CỨU, TÂM SỰ, hoặc câu hỏi thông thường KHÔNG có số tiền, KHÔNG phải nhập liệu mới.\n"
            "   - Output: {\"type\": \"chat\", \"response\": \"Câu trả lời vui vẻ, ngắn gọn, dựa trên Dữ liệu hệ thống (nếu có)...\"}\n"
            "   - Lưu ý: Nếu user hỏi \"tiêu bao nhiêu\", hãy nhìn vào mục 'Hôm nay' hoặc 'Tháng này' trong dữ liệu hệ thống để trả lời chính xác con số.\n"
            "   - Nếu user hỏi câu hỏi thông thường (không liên quan tài chính), hãy trả lời vui vẻ, thân thiện. Nếu cần thông tin thực tế, hãy dùng type \"search\".\n\n"
            "XỬ LÝ NGÀY THÁNG (BACKDATED ENTRY - QUAN TRỌNG):\n"
            f"Thời gian hiện tại: {current_time_str} (Ngày: {current_date_str}).\n"
            "Nhiệm vụ: Trích xuất chi tiêu và NGÀY THÁNG từ input.\n\n"
            "Quy tắc:\n"
            "- Nếu user nói \"Hôm qua\", \"Tối qua\" -> Tính ra ngày hôm qua (so với hiện tại).\n"
            "- Nếu user nói \"Hôm kia\" -> Tính ra ngày hôm kia.\n"
            "- Nếu user nói \"Sáng nay\", \"Tối nay\" -> Dùng ngày hiện tại (date = null).\n"
            "- Nếu user nói \"Ngày 10/12\", \"10/12\" -> Lấy ngày 10/12/{current_time.year}.\n"
            "- Nếu user nói \"Tuần trước\", \"Tháng trước\" -> Tính toán ngày tương ứng.\n"
            "- Nếu không nhắc gì về thời gian -> Mặc định là ngày hiện tại (trả về null hoặc empty).\n\n"
            "Output JSON thêm trường \"date\":\n"
            "{\"type\": \"expense\", \"expenses\": [{\"item\": \"...\", \"amount\": ..., \"category\": \"...\", \"date\": \"DD/MM/YYYY\" hoặc null}]}\n\n"
            "Ví dụ:\n"
            f"  + Input: \"Hôm qua đổ xăng 50k\" (Hôm nay là {current_date_str}) -> Output date: tính ngày hôm qua.\n"
            f"  + Input: \"Ngày 10/12 mua áo 200k\" -> Output date: \"10/12/{current_time.year}\".\n"
            "  + Input: \"Ăn cơm 30k\" (không có thông tin ngày) -> Output date: null.\n\n"
            "XỬ LÝ LỊCH SỬ TRÒ CHUYỆN:\n"
            "- Nếu có lịch sử trò chuyện, hãy tham khảo để hiểu ngữ cảnh.\n"
            "- Khi user hỏi \"chi tiết hơn\", \"tại sao\", \"giải thích\" -> Tham khảo lịch sử để biết user đang hỏi về cái gì.\n"
            "- Hãy trả lời dựa trên ngữ cảnh lịch sử (nếu có) và dữ liệu tài chính.\n\n"
            "QUY TẮC:\n"
            "- Tuyệt đối KHÔNG được nói 'tôi không thể truy cập' hoặc 'tôi không có dữ liệu'.\n"
            "- Dùng dữ liệu hệ thống để trả lời chính xác.\n"
            "- Nếu user hỏi về số liệu, hãy trích xuất số từ dữ liệu hệ thống.\n"
            "- LUÔN trả về JSON chuẩn. Không markdown."
        )
    
    # Debug: Log context data
    if context_data:
        logger.info("=" * 60)
        logger.info("📊 DATA SENT TO AI (CONTEXT):")
        logger.info("=" * 60)
        logger.info(context_data)
        logger.info("=" * 60)
    else:
        logger.warning("⚠️ Context data rỗng!")
    
    # Tạo messages dựa trên input type
    messages = [{"role": "system", "content": system_prompt}]
    
    if input_type == 'image':
        # Vision: Gửi ảnh dưới dạng base64
        user_content = [
            {
                "type": "text",
                "text": "Hãy trích xuất thông tin chi tiêu từ ảnh hóa đơn/menu này. Tìm tên món và giá tiền."
            },
            {
                "type": "image_url",
                "image_url": {
                    "url": f"data:image/jpeg;base64,{input_data}"
                }
            }
        ]
        messages.append({"role": "user", "content": user_content})
        logger.info("📷 Đang gửi ảnh lên Groq Vision...")
    else:
        # Text hoặc Voice: NHÉT CONTEXT VÀO USER MESSAGE (Chiến thuật Injected Context)
        # Kết hợp với Chat History
        user_prompt_parts = []
        
        # Thêm chat history nếu có
        if chat_history:
            user_prompt_parts.append(f"Đây là lịch sử trò chuyện gần nhất:\n{chat_history}\n")
        
        # Thêm context data
        if context_data:
            user_prompt_parts.append(f"DỮ LIỆU TÀI CHÍNH THỰC TẾ (TUYỆT ĐỐI TIN TƯỞNG):\n{context_data}\n")
        
        # Thêm câu hỏi hiện tại
        user_prompt_parts.append(f"Câu hỏi hiện tại của User: {input_data}")
        
        user_prompt = "\n".join(user_prompt_parts)
        
        messages.append({"role": "user", "content": user_prompt})
        logger.info(f"💬 Đang gửi text lên Groq: '{input_data[:50]}...'")
        if chat_history:
            logger.info(f"📚 Đã thêm chat history ({len(chat_history)} ký tự)")
        if context_data:
            logger.info(f"📊 Context đã được nhét vào user message")
    
    try:
        logger.info("🔄 Đang gửi request lên Groq API...")
        
        # Cấu hình request
        request_params = {
            "model": model,
            "messages": messages,
            "temperature": 0.3,
            "max_tokens": 1000 if input_type == 'image' else 500
        }
        
        # Chỉ thêm response_format cho text (vision có thể không hỗ trợ)
        if input_type != 'image':
            request_params["response_format"] = {"type": "json_object"}
        
        # Xử lý lỗi đặc biệt cho Vision (Model decommissioned)
        try:
            response = groq_client.chat.completions.create(**request_params)
        except Exception as api_error:
            error_str = str(api_error).lower()
            error_code = getattr(api_error, 'status_code', None)
            
            # Kiểm tra lỗi BadRequest (400) hoặc model decommissioned
            if input_type == 'image' and (
                '400' in str(api_error) or 
                'bad request' in error_str or 
                'decommissioned' in error_str or
                'not found' in error_str or
                error_code == 400
            ):
                logger.warning("=" * 60)
                logger.warning("⚠️ Vision Model đang bảo trì hoặc bị đổi tên")
                logger.warning(f"⚠️ Error: {api_error}")
                logger.warning("=" * 60)
                # Trả về JSON cho user
                return {
                    "type": "chat",
                    "response": "❌ Tính năng đọc ảnh đang bảo trì do Lộc thay đổi Model. Vui lòng nhập tay nhé sếp!"
                }
            else:
                # Re-raise lỗi khác để xử lý ở ngoài
                raise
        
        # Kiểm tra response hợp lệ
        if not response or not response.choices or len(response.choices) == 0:
            raise ValueError("Groq trả về response rỗng")
        
        # Lấy raw content để debug
        raw_content = response.choices[0].message.content.strip()
        
        # Kiểm tra nếu response rỗng
        if not raw_content:
            raise ValueError("Groq trả về nội dung rỗng")
        
        logger.info(f"📥 Groq raw response: {raw_content}")
        
        # Parse JSON
        try:
            response_data = json.loads(raw_content)
            
            # Kiểm tra format: phải là dict
            if not isinstance(response_data, dict):
                raise ValueError("Groq trả về không phải JSON object")
            
            # Kiểm tra type
            response_type = response_data.get('type', '').lower()
            
            if response_type == 'search':
                # Xử lý search request
                if 'query' not in response_data:
                    raise ValueError("Groq response không có key 'query' cho search")
                
                search_query = response_data['query']
                logger.info(f"✅ Groq AI yêu cầu tìm kiếm: '{search_query}'")
                
                return {
                    'type': 'search',
                    'query': search_query
                }
            elif response_type == 'qr_request':
                # Xử lý QR request
                if 'amount' not in response_data:
                    raise ValueError("Groq response không có key 'amount' cho qr_request")
                
                amount = response_data['amount']
                content = response_data.get('content', '')
                
                # Validate amount
                if not isinstance(amount, int) or amount <= 0:
                    raise ValueError(f"Amount không hợp lệ: {amount}")
                
                logger.info(f"✅ Groq AI yêu cầu tạo QR: {amount:,}đ - '{content}'")
                
                return {
                    'type': 'qr_request',
                    'amount': amount,
                    'content': content
                }
            elif response_type == 'expense':
                # Xử lý chi tiêu
                if 'expenses' not in response_data:
                    raise ValueError("Groq response không có key 'expenses'")
                
                expenses_data = response_data['expenses']
                
                # Đảm bảo là list
                if not isinstance(expenses_data, list):
                    expenses_data = [expenses_data] if expenses_data else []
                
                # Validate và format kết quả
                results = []
                for item in expenses_data:
                    if not isinstance(item, dict):
                        continue
                    
                    # Validate required fields
                    if 'item' not in item or 'amount' not in item:
                        logger.warning(f"⚠️ Item thiếu field: {item}")
                        continue
                    
                    # Validate category
                    category = item.get('category', 'Khác')
                    if category not in ['Ăn uống', 'Di chuyển', 'Học tập', 'Khác']:
                        category = 'Khác'
                    
                    # Validate amount (phải là số)
                    try:
                        amount = int(item['amount'])
                        if amount <= 0:
                            logger.warning(f"⚠️ Amount <= 0: {amount}")
                            continue
                    except (ValueError, TypeError):
                        logger.warning(f"⚠️ Amount không hợp lệ: {item.get('amount')}")
                        continue
                    
                    # Xử lý item name
                    item_name = str(item['item']).strip()
                    if not item_name or item_name == "Chưa rõ":
                        item_name = "Không xác định"
                    
                    # Xử lý date (backdated entry)
                    expense_date = item.get('date')
                    if expense_date:
                        # Validate format DD/MM/YYYY
                        try:
                            # Kiểm tra format
                            if isinstance(expense_date, str) and '/' in expense_date:
                                # Giữ nguyên date string để parse sau
                                logger.info(f"📅 Expense có date: {expense_date}")
                            else:
                                expense_date = None
                        except:
                            expense_date = None
                    else:
                        expense_date = None
                    
                    expense_dict = {
                        'item': item_name,
                        'amount': amount,
                        'category': category
                    }
                    if expense_date:
                        expense_dict['date'] = expense_date
                    
                    results.append(expense_dict)
                
                if not results:
                    raise ValueError("Groq trả về list expenses rỗng")
                
                # Lấy message từ AI (nếu có)
                ai_message = response_data.get('message', '')
                
                logger.info(f"✅ Groq AI đã phân tích thành công {len(results)} món")
                for i, expense in enumerate(results, 1):
                    logger.info(f"  {i}. {expense['item']}: {expense['amount']:,}đ ({expense['category']})")
                if ai_message:
                    logger.info(f"💬 AI message: {ai_message}")
                logger.info("=" * 60)
                
                return {
                    'type': 'expense',
                    'expenses': results,
                    'message': ai_message
                }
                
            elif response_type == 'chat':
                # Xử lý chat
                if 'response' not in response_data:
                    raise ValueError("Groq response không có key 'response'")
                
                chat_response = response_data.get('response', '')
                
                if not chat_response:
                    raise ValueError("Groq trả về response rỗng")
                
                logger.info(f"✅ Groq AI đã phân tích: Chat mode")
                logger.info(f"💬 AI response: {chat_response}")
                logger.info("=" * 60)
                
                return {
                    'type': 'chat',
                    'response': chat_response
                }
            else:
                raise ValueError(f"Groq trả về type không hợp lệ: {response_type}")
            
        except json.JSONDecodeError as e:
            logger.error(f"❌ Lỗi parse JSON từ Groq: {e}")
            logger.error(f"📝 Raw response: {raw_content}")
            raise ValueError(f"Groq trả về JSON không hợp lệ: {e}")
            
    except Exception as e:
        error_str = str(e).lower()
        # Kiểm tra các loại lỗi phổ biến
        if 'quota' in error_str or 'rate limit' in error_str or '429' in error_str:
            logger.warning("=" * 60)
            logger.warning("⚠️ GROQ API QUOTA ĐÃ HẾT")
            logger.warning("=" * 60)
            logger.warning("💡 Bot sẽ tự động chuyển sang Regex Parsing")
            logger.warning("💡 Kiểm tra quota tại: https://console.groq.com/usage")
            logger.warning("=" * 60)
        elif 'api key' in error_str or '401' in error_str or '403' in error_str:
            logger.warning("⚠️ Lỗi xác thực Groq API (API Key không hợp lệ)")
        else:
            logger.error(f"❌ Lỗi khi gọi Groq API: {e}")
        raise


def parse_multiple_items(text: str) -> list:
    """
    Parse nhiều món từ một tin nhắn (Regex Fallback)
    Hỗ trợ phân cách bởi dấu phẩy hoặc xuống dòng
    """
    logger.info("=" * 60)
    logger.info("🔍 ĐANG SỬ DỤNG REGEX PARSING (Fallback)")
    logger.info("=" * 60)
    logger.info(f"📝 Text nhận được: '{text}'")
    
    # Tách text thành các phần (dấu phẩy hoặc xuống dòng)
    # Loại bỏ khoảng trắng thừa
    text = text.strip()
    
    # Tách theo dấu phẩy hoặc xuống dòng
    items_text = re.split(r'[,，\n\r]+', text)
    items_text = [item.strip() for item in items_text if item.strip()]
    
    logger.info(f"🔍 Đã tách thành {len(items_text)} phần")
    
    results = []
    for i, item_text in enumerate(items_text, 1):
        logger.info(f"🔍 Đang xử lý món {i}/{len(items_text)}: '{item_text}'")
        try:
            parsed_item = parse_single_item(item_text)
            results.append(parsed_item)
            logger.info(f"✅ Món {i}: {parsed_item['item']} - {parsed_item['amount']:,}đ - {parsed_item['category']}")
        except ValueError as e:
            logger.warning(f"⚠️ Món {i} không hợp lệ: {e}")
            continue
    
    if not results:
        raise ValueError("Không tìm thấy món hợp lệ nào trong tin nhắn")
    
    logger.info("=" * 60)
    logger.info(f"✅ Đã phân tích thành công {len(results)} món")
    logger.info("=" * 60)
    
    return results

# ==================== KẾT NỐI GOOGLE SHEETS ====================
# Sử dụng hàm từ services.py
try:
    worksheet = init_google_sheets()
except Exception as e:
    logger.critical("❌ KHÔNG THỂ KHỞI ĐỘNG BOT!")
    raise

# ==================== LƯU VÀO GOOGLE SHEET ====================
# Hàm đã được tách ra services.py, import ở trên
# Không cần định nghĩa lại, sử dụng trực tiếp từ services

# ==================== HOÀN TÁC (UNDO) ====================
def undo_last_expense() -> dict:
    """
    Xóa dòng cuối cùng có dữ liệu trong Google Sheet
    Trả về thông tin dòng đã xóa hoặc None nếu không có gì để xóa
    """
    logger.info("=" * 60)
    logger.info("BƯỚC: HOÀN TÁC GIAO DỊCH CUỐI")
    logger.info("=" * 60)
    
    if worksheet is None:
        raise ValueError("Google Sheets chưa được khởi tạo")
    
    try:
        all_data = worksheet.get_all_values()
        
        # Kiểm tra nếu Sheet trống hoặc chỉ có header
        if len(all_data) <= 1:
            logger.warning("⚠️ Sheet trống, không có gì để xóa")
            return None
        
        # Lấy dòng cuối cùng (bỏ qua header)
        last_row_index = len(all_data)
        last_row = all_data[-1]
        
        # Kiểm tra xem dòng có dữ liệu không
        if len(last_row) < 7 or not last_row[4]:  # Cột 5 (index 4) là Tên món
            logger.warning("⚠️ Dòng cuối không có dữ liệu hợp lệ")
            return None
        
        # Lấy thông tin dòng sẽ xóa
        deleted_info = {
            'item': last_row[4] if len(last_row) > 4 else 'Không xác định',
            'amount': int(last_row[6]) if len(last_row) > 6 and last_row[6] else 0,
            'category': last_row[5] if len(last_row) > 5 else 'Khác',
            'date': f"{last_row[1]}/{last_row[2]}/{last_row[3]}" if len(last_row) > 3 else 'N/A'
        }
        
        # Xóa dòng cuối cùng
        logger.info(f"🗑️ Đang xóa dòng {last_row_index}: {deleted_info['item']} - {deleted_info['amount']:,}đ")
        worksheet.delete_rows(last_row_index)
        
        logger.info("=" * 60)
        logger.info("✅ Đã xóa giao dịch cuối cùng thành công!")
        logger.info("=" * 60)
        
        return deleted_info
        
    except Exception as e:
        logger.error(f"❌ Lỗi khi xóa giao dịch: {e}")
        raise


def get_wasteful_warning(item_name: str) -> str:
    """
    Kiểm tra xem tên món có chứa từ khóa lãng phí không
    Nếu có, trả về một câu cảnh báo ngẫu nhiên
    """
    item_lower = item_name.lower()
    
    for keyword in WASTEFUL_KEYWORDS:
        if keyword in item_lower:
            import random
            warning = random.choice(WASTEFUL_WARNINGS)
            logger.info(f"⚠️ Phát hiện từ khóa lãng phí: '{keyword}' trong '{item_name}'")
            return warning
    
    return None


# ==================== TELEGRAM HANDLERS ====================
async def start_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Xử lý lệnh /start"""
    logger.info(f"📨 Nhận lệnh /start từ user: {update.effective_user.id}")
    welcome_message = (
        "Chào bạn! 👋\n\n"
        "🤖 **Bot Quản Lý Chi Tiêu Enterprise Edition**\n\n"
        "📝 **Cách sử dụng:**\n"
        "• Gửi một món: `phở 50k`\n"
        "• Gửi nhiều món: `cơm 35k, trà đá 5k, xăng 50k`\n"
        "• Hoặc xuống dòng:\n"
        "  `phở 50k`\n"
        "  `cơm 35k`\n\n"
        "💡 Gõ `/help` để xem hướng dẫn đầy đủ!\n\n"
        "Hỗ trợ: k, ng, nghìn, tr, triệu, d, đ"
    )
    await update.message.reply_text(welcome_message, parse_mode=ParseMode.MARKDOWN)
    logger.info("✅ Đã gửi welcome message")


async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Xử lý lệnh /help - Hiển thị hướng dẫn đầy đủ"""
    logger.info(f"📨 Nhận lệnh /help từ user: {update.effective_user.id}")
    
    help_message = (
        "📚 **HƯỚNG DẪN SỬ DỤNG BOT**\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
        
        "📝 **1. THÊM CHI TIÊU**\n"
        "Gửi tin nhắn mô tả chi tiêu:\n"
        "• `phở 50k` - Một món\n"
        "• `cơm 35k, trà đá 5k, xăng 50k` - Nhiều món (phân cách bằng dấu phẩy)\n"
        "• Hoặc xuống dòng:\n"
        "  `phở 50k`\n"
        "  `cơm 35k`\n\n"
        
        "💡 **Định dạng số tiền hỗ trợ:**\n"
        "• `35k`, `50ng`, `30 nghìn` → 35,000đ\n"
        "• `1.5tr`, `2 triệu` → 1,500,000đ\n"
        "• `50000`, `50000đ`, `50000d` → 50,000đ\n\n"
        
        "📊 **2. BÁO CÁO & THỐNG KÊ**\n"
        "• `/report` hoặc `/thongke`\n"
        "  → Xem báo cáo chi tiêu hôm nay, tháng này, top chi tiêu\n\n"
        
        "• `/chart`\n"
        "  → Xem biểu đồ tròn (Donut Chart) chi tiêu tháng này\n"
        "  → Hiển thị tỷ lệ % theo từng phân loại\n\n"
        
        "• `/export`\n"
        "  → Xuất báo cáo Excel tháng này\n"
        "  → File Excel chuyên nghiệp, có format đẹp\n\n"
        
        "💳 **3. MÃ QR CHUYỂN KHOẢN**\n"
        "• `/pay 50k trả tiền cơm`\n"
        "  → Tạo mã QR chuyển khoản nhanh VietQR\n"
        "  → Quét mã để chuyển tiền cho sếp Lộc\n"
        "• `/qr 100k tiền cafe` - Alias cho /pay\n\n"
        
        "🔧 **4. QUẢN LÝ**\n"
        "• `/undo`\n"
        "  → Hoàn tác giao dịch cuối cùng\n"
        "  → Xóa dòng cuối cùng trong Sheet\n\n"
        
        "🔔 **4. BÁO THỨC NHẬP LIỆU**\n"
        "• `/remind 21:30`\n"
        "  → Đặt báo thức nhắc nhở hàng ngày lúc 21:30\n"
        "  → Bot sẽ tự động nhắc bạn tổng kết chi tiêu\n\n"
        
        "• `/stopremind`\n"
        "  → Tắt báo thức nhắc nhở\n\n"
        
        "🧾 **5. MÁY TÍNH CHIA TIỀN**\n"
        "• `/chia 500k 4`\n"
        "  → Chia 500.000đ cho 4 người\n"
        "  → Kết quả: Mỗi người 125.000đ\n\n"
        
        "• `/chia 300k Nam, Hùng, Lộc`\n"
        "  → Chia 300.000đ cho 3 người\n"
        "  → Hiển thị chi tiết từng người\n\n"
        
        "💰 **6. QUẢN LÝ NGÂN SÁCH TUẦN**\n"
        "• Hạn mức: **700,000đ/tuần**\n"
        "• Bot tự động theo dõi và cảnh báo:\n"
        "  → Hiển thị số dư còn lại sau mỗi giao dịch\n"
        "  → Cảnh báo nếu tiêu quá 80% và mới đầu tuần\n"
        "  → Báo động nếu vượt quá hạn mức\n\n"
        
        "🚨 **7. CẢNH SÁT CHI TIÊU**\n"
        "Bot tự động phát hiện và cảnh báo các khoản chi lãng phí:\n"
        "• Game: nạp, skin, gacha, top up...\n"
        "• Đồ uống: trà sữa, toco, mixue...\n"
        "• Giải trí: phim, netflix...\n"
        "• Khác: đồ chơi, mô hình, nhậu...\n\n"
        
        "🏷️ **8. PHÂN LOẠI TỰ ĐỘNG**\n"
        "Bot tự động phân loại dựa trên từ khóa:\n"
        "• **Ăn uống:** phở, cơm, bún, cafe, trà...\n"
        "• **Di chuyển:** xăng, xe, grab, taxi...\n"
        "• **Học tập:** sách, vở, bút, học phí...\n"
        "• **Khác:** Nếu không khớp từ khóa nào\n\n"
        
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
        "💡 **Mẹo sử dụng:**\n"
        "• Gõ `/help` để xem lại hướng dẫn này\n"
        "• Gõ `/start` để xem lời chào\n"
        "• Bot hoạt động offline, không cần AI\n"
        "• Tất cả dữ liệu được lưu vào Google Sheet\n\n"
        
        "🎯 **Phiên bản: Enterprise Edition**"
    )
    
    await update.message.reply_text(help_message, parse_mode=ParseMode.MARKDOWN)
    logger.info("✅ Đã gửi hướng dẫn cho user")


async def report_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Xử lý lệnh /report hoặc /thongke"""
    logger.info(f"📨 Nhận lệnh /report từ user: {update.effective_user.id}")
    
    try:
        report_data = get_expense_report()
        
        today_total = report_data['today_total']
        month_total = report_data['month_total']
        top_expenses = report_data['top_expenses']
        
        now = datetime.now()
        month_name = now.strftime('%B')
        
        # Tạo message báo cáo
        report_message = f"📊 **BÁO CÁO CHI TIÊU**\n"
        report_message += "━━━━━━━━━━━━━━━━━━\n"
        report_message += f"📅 Hôm nay: **{today_total:,}đ**\n"
        report_message += f"🗓️ Tháng {now.month}: **{month_total:,}đ**\n"
        report_message += "━━━━━━━━━━━━━━━━━━\n"
        
        if top_expenses:
            report_message += "🔥 **Top chi tiêu tháng:**\n"
            for i, (category, amount) in enumerate(top_expenses, 1):
                report_message += f"{i}. {category}: {amount:,}đ\n"
        else:
            report_message += "📝 Chưa có dữ liệu chi tiêu trong tháng này.\n"
        
        await update.message.reply_text(report_message, parse_mode=ParseMode.MARKDOWN)
        logger.info("✅ Đã gửi báo cáo cho user")
        
    except Exception as e:
        logger.error(f"❌ Lỗi khi tạo báo cáo: {e}")
        error_msg = "❌ Đã xảy ra lỗi khi tạo báo cáo. Vui lòng thử lại sau."
        await update.message.reply_text(error_msg)


async def undo_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Xử lý lệnh /undo - Xóa giao dịch cuối cùng"""
    logger.info(f"📨 Nhận lệnh /undo từ user: {update.effective_user.id}")
    
    try:
        deleted_info = undo_last_expense()
        
        if deleted_info is None:
            response = "❌ Không có gì để xóa.\n\nSheet trống hoặc không có giao dịch nào."
        else:
            response = f"✅ **Đã xóa giao dịch cuối cùng thành công!**\n\n"
            response += f"📝 Giao dịch đã xóa:\n"
            response += f"• {deleted_info['item']}: {deleted_info['amount']:,}đ\n"
            response += f"• Phân loại: {deleted_info['category']}\n"
            response += f"• Ngày: {deleted_info['date']}"
        
        await update.message.reply_text(response, parse_mode=ParseMode.MARKDOWN)
        logger.info("✅ Đã gửi phản hồi undo cho user")
        
    except Exception as e:
        logger.error(f"❌ Lỗi khi xóa giao dịch: {e}")
        error_msg = "❌ Đã xảy ra lỗi khi xóa giao dịch. Vui lòng thử lại sau."
        await update.message.reply_text(error_msg)


async def delete_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Xử lý lệnh /delete hoặc /xoa - Xóa giao dịch theo tên món (tìm kiếm thông minh)"""
    logger.info(f"📨 Nhận lệnh /delete từ user: {update.effective_user.id}")
    
    try:
        # Kiểm tra xem user có nhập tên món không
        if not context.args or len(context.args) == 0:
            response = (
                "❌ **Sai cú pháp!**\n\n"
                "💡 Cách sử dụng:\n"
                "• `/delete com ga` - Xóa món 'Cơm gà' (tìm trong hôm nay)\n"
                "• `/delete bun bo` - Xóa món 'Bún bò' (tìm trong hôm nay)\n"
                "• `/xoa pho` - Xóa món 'Phở' (tìm trong hôm nay)\n\n"
                "🤖 Bot sẽ tự động tìm món tương đồng nếu bạn gõ không chính xác 100%.\n"
                "Ví dụ: Gõ 'com ga' sẽ tìm thấy 'Cơm gà xối mỡ'."
            )
            await update.message.reply_text(response, parse_mode=ParseMode.MARKDOWN)
            return
        
        # Lấy từ khóa tìm kiếm từ user
        user_input = ' '.join(context.args)
        logger.info(f"🔍 User muốn xóa món: '{user_input}'")
        
        # Tìm kiếm trong hôm nay (có thể mở rộng để tìm trong tháng)
        search_result = find_expense_by_name(user_input, search_in_month=False)
        
        if not search_result['found']:
            # Không tìm thấy, thử tìm trong tháng này
            search_result = find_expense_by_name(user_input, search_in_month=True)
            
            if not search_result['found']:
                response = f"❌ Không tìm thấy món nào tên giống '{user_input}' cả.\n\n"
                response += "💡 Hãy thử:\n"
                response += "• Gõ tên món chính xác hơn\n"
                response += "• Kiểm tra lại xem món đã được thêm vào chưa"
                await update.message.reply_text(response, parse_mode=ParseMode.MARKDOWN)
                return
        
        # Tìm thấy 1 món tương đồng
        match = search_result['match']
        row_index = search_result['row_index']
        
        # Tạo Inline Keyboard để xác nhận
        keyboard = [
            [
                InlineKeyboardButton("✅ Đúng, xóa đi", callback_data=f"delete_confirm_{row_index}"),
                InlineKeyboardButton("❌ Không phải", callback_data="delete_cancel")
            ]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        # Gửi câu hỏi xác nhận
        response = (
            f"🔍 **Tìm thấy món tương đồng:**\n\n"
            f"📝 **{match['item']}**\n"
            f"💰 {match['amount']:,}đ\n"
            f"📂 {match['category']}\n"
            f"📅 {match['date']}\n\n"
            f"❓ Có phải bạn muốn xóa món này không?"
        )
        
        await update.message.reply_text(response, parse_mode=ParseMode.MARKDOWN, reply_markup=reply_markup)
        logger.info(f"✅ Đã gửi câu hỏi xác nhận cho user (row_index: {row_index})")
        
    except Exception as e:
        logger.error(f"❌ Lỗi khi xử lý lệnh delete: {e}", exc_info=True)
        error_msg = "❌ Đã xảy ra lỗi khi tìm kiếm giao dịch. Vui lòng thử lại sau."
        await update.message.reply_text(error_msg)


async def delete_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Xử lý callback từ Inline Keyboard khi user xác nhận xóa"""
    query = update.callback_query
    await query.answer()
    
    logger.info(f"📨 Nhận callback delete từ user: {update.effective_user.id}")
    
    try:
        callback_data = query.data
        
        if callback_data == "delete_cancel":
            # User bấm "Không phải"
            await query.edit_message_text("❌ Đã hủy xóa giao dịch.")
            logger.info("✅ User đã hủy xóa")
            return
        
        if callback_data.startswith("delete_confirm_"):
            # User bấm "Đúng, xóa đi"
            row_index = int(callback_data.split("_")[2])
            logger.info(f"🗑️ User xác nhận xóa dòng {row_index}")
            
            # Xóa giao dịch
            deleted_info = delete_expense_by_row_index(row_index)
            
            # Cập nhật message
            response = (
                f"✅ **Đã xóa giao dịch thành công!**\n\n"
                f"📝 Giao dịch đã xóa:\n"
                f"• **{deleted_info['item']}**: {deleted_info['amount']:,}đ\n"
                f"• Phân loại: {deleted_info['category']}\n"
                f"• Ngày: {deleted_info['date']}"
            )
            
            await query.edit_message_text(response, parse_mode=ParseMode.MARKDOWN)
            logger.info("✅ Đã xóa giao dịch và cập nhật message")
            
    except Exception as e:
        logger.error(f"❌ Lỗi khi xử lý callback delete: {e}", exc_info=True)
        error_msg = "❌ Đã xảy ra lỗi khi xóa giao dịch. Vui lòng thử lại sau."
        try:
            await query.edit_message_text(error_msg)
        except:
            await query.message.reply_text(error_msg)


# ==================== BÁO THỨC NHẬP LIỆU ====================
async def remind_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Xử lý lệnh /remind - Đặt báo thức nhắc nhở hàng ngày"""
    logger.info(f"📨 Nhận lệnh /remind từ user: {update.effective_user.id}")
    user_id = str(update.effective_user.id)
    
    try:
        if not context.args or len(context.args) == 0:
            response = (
                "❌ **Sai cú pháp!**\n\n"
                "💡 Cách sử dụng:\n"
                "• `/remind 21:30` - Đặt báo thức lúc 21:30 hàng ngày\n"
                "• `/remind 09:00` - Đặt báo thức lúc 9:00 sáng\n\n"
                "Ví dụ: `/remind 21:30`"
            )
            await update.message.reply_text(response, parse_mode=ParseMode.MARKDOWN)
            return
        
        time_str = context.args[0]
        
        # Parse thời gian (HH:MM)
        try:
            time_parts = time_str.split(':')
            if len(time_parts) != 2:
                raise ValueError("Sai định dạng")
            
            hour = int(time_parts[0])
            minute = int(time_parts[1])
            
            if hour < 0 or hour > 23 or minute < 0 or minute > 59:
                raise ValueError("Giờ không hợp lệ")
            
            # Lưu reminder
            user_reminders[user_id] = {'hour': hour, 'minute': minute}
            save_reminders()
            
            # Lên lịch job
            job_queue = context.application.job_queue
            if job_queue:
                # Xóa job cũ nếu có
                current_jobs = job_queue.get_jobs_by_name(f"reminder_{user_id}")
                for job in current_jobs:
                    job.schedule_removal()
                
                # Tạo job mới - chạy hàng ngày vào giờ đã đặt
                reminder_time = dt_time(hour, minute)
                job_queue.run_daily(
                    send_daily_reminder,
                    time=reminder_time,
                    name=f"reminder_{user_id}",
                    chat_id=update.effective_chat.id
                )
                
                # Lưu chat_id vào reminder data để khôi phục sau khi restart
                user_reminders[user_id]['chat_id'] = update.effective_chat.id
                save_reminders()
            
            response = (
                f"✅ **Đã đặt báo thức thành công!**\n\n"
                f"🔔 Bot sẽ nhắc bạn hàng ngày lúc **{hour:02d}:{minute:02d}**\n\n"
                f"💡 Gõ `/stopremind` để tắt báo thức"
            )
            await update.message.reply_text(response, parse_mode=ParseMode.MARKDOWN)
            logger.info(f"✅ Đã đặt reminder cho user {user_id} lúc {hour:02d}:{minute:02d}")
            
        except (ValueError, IndexError) as e:
            response = (
                "❌ **Sai định dạng giờ!**\n\n"
                "💡 Định dạng đúng: `HH:MM`\n"
                "• Ví dụ: `21:30`, `09:00`, `18:45`\n"
                "• Giờ: 00-23, Phút: 00-59"
            )
            await update.message.reply_text(response, parse_mode=ParseMode.MARKDOWN)
            
    except Exception as e:
        logger.error(f"❌ Lỗi khi đặt reminder: {e}")
        error_msg = "❌ Đã xảy ra lỗi. Vui lòng thử lại sau."
        await update.message.reply_text(error_msg)


async def stopremind_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Xử lý lệnh /stopremind - Tắt báo thức"""
    logger.info(f"📨 Nhận lệnh /stopremind từ user: {update.effective_user.id}")
    user_id = str(update.effective_user.id)
    
    try:
        if user_id in user_reminders:
            # Xóa reminder
            del user_reminders[user_id]
            save_reminders()
            
            # Xóa job
            job_queue = context.application.job_queue
            if job_queue:
                current_jobs = job_queue.get_jobs_by_name(f"reminder_{user_id}")
                for job in current_jobs:
                    job.schedule_removal()
            
            response = "✅ **Đã tắt báo thức nhắc nhở!**\n\n💡 Gõ `/remind [giờ]` để đặt lại"
            logger.info(f"✅ Đã tắt reminder cho user {user_id}")
        else:
            response = "ℹ️ Bạn chưa đặt báo thức nào.\n\n💡 Gõ `/remind [giờ]` để đặt báo thức"
        
        await update.message.reply_text(response, parse_mode=ParseMode.MARKDOWN)
        
    except Exception as e:
        logger.error(f"❌ Lỗi khi tắt reminder: {e}")
        error_msg = "❌ Đã xảy ra lỗi. Vui lòng thử lại sau."
        await update.message.reply_text(error_msg)


async def send_daily_reminder(context: ContextTypes.DEFAULT_TYPE):
    """Gửi tin nhắn nhắc nhở hàng ngày"""
    chat_id = context.job.chat_id
    reminder_message = (
        "🔔 **Nhắc nhở:**\n\n"
        "Đừng quên tổng kết chi tiêu hôm nay nhé! 💸\n\n"
        "💡 Gõ `/report` để xem báo cáo chi tiêu"
    )
    
    try:
        await context.bot.send_message(
            chat_id=chat_id,
            text=reminder_message,
            parse_mode=ParseMode.MARKDOWN
        )
        logger.info(f"✅ Đã gửi reminder cho chat {chat_id}")
    except Exception as e:
        logger.error(f"❌ Lỗi khi gửi reminder: {e}")


# ==================== MÁY TÍNH CHIA TIỀN ====================
def parse_amount_for_split(text: str) -> int:
    """Parse số tiền từ text (dùng cho bill splitter)"""
    text_lower = text.lower().strip()
    
    patterns = [
        (r'(\d+(?:\.\d+)?)\s*tr(?:iệu)?', 1000000),
        (r'(\d+(?:\.\d+)?)\s*k(?:ilo)?', 1000),
        (r'(\d+(?:\.\d+)?)\s*ng(?:àn)?', 1000),
        (r'(\d+(?:\.\d+)?)\s*nghìn', 1000),
        (r'(\d+(?:\.\d+)?)\s*000', 1),
        (r'(\d+(?:\.\d+)?)\s*d(?:ồng)?', 1),
        (r'(\d+(?:\.\d+)?)\s*đ', 1),
        (r'(\d{4,})', 1),
    ]
    
    for pattern, multiplier in patterns:
        match = re.search(pattern, text_lower)
        if match:
            try:
                number = float(match.group(1))
                return int(number * multiplier)
            except:
                continue
    
    return 0


async def chia_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Xử lý lệnh /chia - Chia tiền giữa nhiều người"""
    logger.info(f"📨 Nhận lệnh /chia từ user: {update.effective_user.id}")
    
    try:
        if not context.args or len(context.args) < 2:
            response = (
                "❌ **Sai cú pháp!**\n\n"
                "💡 **Cách sử dụng:**\n"
                "• `/chia 500k 4` - Chia 500k cho 4 người\n"
                "• `/chia 300k Nam, Hùng, Lộc` - Chia 300k cho 3 người\n\n"
                "**Ví dụ:**\n"
                "• `/chia 500k 4`\n"
                "• `/chia 1tr Nam, Hùng, Lộc, Mai`"
            )
            await update.message.reply_text(response, parse_mode=ParseMode.MARKDOWN)
            return
        
        # Parse số tiền
        amount_text = context.args[0]
        total_amount = parse_amount_for_split(amount_text)
        
        if total_amount == 0:
            response = (
                "❌ **Không tìm thấy số tiền hợp lệ!**\n\n"
                "💡 Định dạng số tiền:\n"
                "• `500k`, `1tr`, `500000`, `500000đ`"
            )
            await update.message.reply_text(response, parse_mode=ParseMode.MARKDOWN)
            return
        
        # Parse số người hoặc danh sách tên
        remaining_args = ' '.join(context.args[1:])
        
        # Kiểm tra xem có phải là số không
        try:
            num_people = int(remaining_args)
            # Trường hợp 1: Chia cho số người
            if num_people <= 0:
                raise ValueError("Số người phải > 0")
            
            per_person = total_amount // num_people
            remainder = total_amount % num_people
            
            response = f"🧾 **HÓA ĐƠN CHIA TIỀN**\n"
            response += f"💰 Tổng: {total_amount:,}đ\n"
            response += f"👥 Số người: {num_people}\n"
            response += "━━━━━━━━━━━━━━━━━━\n"
            response += f"💵 **Mỗi người: {per_person:,}đ**\n"
            
            if remainder > 0:
                response += f"⚠️ Dư: {remainder:,}đ (có thể để tiền lẻ hoặc ai đó chịu thêm)\n"
            
            response += "━━━━━━━━━━━━━━━━━━\n"
            response += "👉 *Copy đoạn này gửi đòi nợ nhé!*"
            
        except ValueError:
            # Trường hợp 2: Chia theo danh sách tên
            # Tách tên bằng dấu phẩy
            names = [name.strip() for name in remaining_args.split(',')]
            names = [name for name in names if name]  # Loại bỏ tên rỗng
            
            if len(names) == 0:
                response = (
                    "❌ **Không tìm thấy tên người!**\n\n"
                    "💡 Ví dụ:\n"
                    "• `/chia 300k Nam, Hùng, Lộc`\n"
                    "• `/chia 500k An, Bình, Chi, Dung`"
                )
                await update.message.reply_text(response, parse_mode=ParseMode.MARKDOWN)
                return
            
            num_people = len(names)
            per_person = total_amount // num_people
            remainder = total_amount % num_people
            
            response = f"🧾 **HÓA ĐƠN CHIA TIỀN**\n"
            response += f"💰 Tổng: {total_amount:,}đ\n"
            response += f"👥 Số người: {num_people}\n"
            response += "━━━━━━━━━━━━━━━━━━\n"
            
            # Hiển thị từng người
            for i, name in enumerate(names):
                amount_for_person = per_person
                # Người cuối cùng nhận phần dư (nếu có)
                if i == len(names) - 1 and remainder > 0:
                    amount_for_person += remainder
                    response += f"👤 **{name}**: {amount_for_person:,}đ (gồm {remainder:,}đ dư)\n"
                else:
                    response += f"👤 **{name}**: {amount_for_person:,}đ\n"
            
            response += "━━━━━━━━━━━━━━━━━━\n"
            response += "👉 *Copy đoạn này gửi đòi nợ nhé!*"
        
        await update.message.reply_text(response, parse_mode=ParseMode.MARKDOWN)
        logger.info(f"✅ Đã tính chia tiền: {total_amount:,}đ cho {num_people} người")
        
    except Exception as e:
        logger.error(f"❌ Lỗi khi chia tiền: {e}", exc_info=True)
        error_msg = "❌ Đã xảy ra lỗi. Vui lòng thử lại sau."
        await update.message.reply_text(error_msg)


# ==================== BIỂU ĐỒ TRỰC QUAN ====================
def get_monthly_data() -> pd.DataFrame:
    """Đọc dữ liệu tháng hiện tại từ Sheet và trả về DataFrame"""
    logger.info("=" * 60)
    logger.info("BƯỚC: ĐỌC DỮ LIỆU THÁNG HIỆN TẠI")
    logger.info("=" * 60)
    
    if worksheet is None:
        raise ValueError("Google Sheets chưa được khởi tạo")
    
    try:
        all_data = worksheet.get_all_values()
        if len(all_data) <= 1:  # Chỉ có header
            return pd.DataFrame()
        
        data_rows = all_data[1:]
        now = datetime.now()
        current_month = now.month
        current_year = now.year
        
        # Lọc dữ liệu tháng này
        monthly_data = []
        for row in data_rows:
            if len(row) < 7:
                continue
            
            try:
                row_day = int(row[1]) if row[1] else 0
                row_month = int(row[2]) if row[2] else 0
                row_year = int(row[3]) if row[3] else 0
                
                if row_month == current_month and row_year == current_year:
                    monthly_data.append({
                        'Full Time': row[0] if len(row) > 0 else '',
                        'Ngày': row_day,
                        'Tháng': row_month,
                        'Năm': row_year,
                        'Tên món': row[4] if len(row) > 4 else 'Không xác định',
                        'Phân loại': row[5] if len(row) > 5 else 'Khác',
                        'Số tiền': int(row[6]) if row[6] else 0
                    })
            except (ValueError, IndexError):
                continue
        
        df = pd.DataFrame(monthly_data)
        logger.info(f"✅ Đã đọc {len(df)} dòng dữ liệu tháng {current_month}/{current_year}")
        return df
        
    except Exception as e:
        logger.error(f"❌ Lỗi khi đọc dữ liệu: {e}")
        raise


async def pay_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Xử lý lệnh /pay hoặc /qr - Tạo mã QR chuyển khoản nhanh VietQR"""
    logger.info(f"📨 Nhận lệnh /pay từ user: {update.effective_user.id}")
    
    try:
        if not context.args or len(context.args) < 1:
            response = (
                "⚠️ **Sai cú pháp!**\n\n"
                "💡 **Cách sử dụng:**\n"
                "• `/pay 50k trả tiền cơm`\n"
                "• `/pay 100k tiền cafe`\n"
                "• `/pay 500k` (không có nội dung)\n\n"
                "**Ví dụ:**\n"
                "• `/pay 50k trả tiền cơm`\n"
                "• `/pay 1tr tiền nhà`"
            )
            await update.message.reply_text(response, parse_mode=ParseMode.MARKDOWN)
            return
        
        # Parse số tiền từ argument đầu tiên
        amount_text = context.args[0]
        
        # Sử dụng hàm parse_amount_for_split để parse số tiền
        amount = parse_amount_for_split(amount_text)
        
        if amount == 0:
            response = (
                "❌ **Không tìm thấy số tiền hợp lệ!**\n\n"
                "💡 Định dạng số tiền:\n"
                "• `50k`, `100ng`, `500000`, `1tr`"
            )
            await update.message.reply_text(response, parse_mode=ParseMode.MARKDOWN)
            return
        
        # Lấy nội dung chuyển khoản (tất cả arguments còn lại)
        content = " ".join(context.args[1:]) if len(context.args) > 1 else ""
        
        logger.info(f"💰 Số tiền: {amount:,}đ")
        logger.info(f"📝 Nội dung: '{content}'")
        
        # Tạo URL VietQR
        qr_url = generate_vietqr_url(amount, content)
        
        if not qr_url:
            await update.message.reply_text(
                "❌ Không thể tạo mã QR. Vui lòng thử lại sau."
            )
            return
        
        # Tải ảnh QR từ URL
        try:
            img_response = requests.get(qr_url, timeout=10)
            if img_response.status_code == 200:
                image_buffer = io.BytesIO(img_response.content)
                image_buffer.seek(0)
                
                # Tạo caption
                caption = (
                    f"💳 **Quét mã này bank cho sếp Lộc nha!**\n"
                    f"💰 **Số tiền:** {amount:,}đ\n"
                    f"🏦 **VPBank - 0375646013**\n"
                    f"👤 **LE PHUOC LOC**"
                )
                if content:
                    caption += f"\n📝 **Nội dung:** {content}"
                
                # Gửi ảnh QR code
                await update.message.reply_photo(
                    photo=image_buffer,
                    caption=caption,
                    parse_mode=ParseMode.MARKDOWN
                )
                logger.info("✅ Đã gửi mã QR VietQR cho user")
            else:
                await update.message.reply_text(
                    f"❌ Không thể tải ảnh QR (HTTP {img_response.status_code})"
                )
        except Exception as e:
            logger.error(f"❌ Lỗi tải ảnh QR: {e}", exc_info=True)
            await update.message.reply_text(
                "❌ Không thể tải ảnh QR. Vui lòng thử lại sau."
            )
        
    except Exception as e:
        logger.error(f"❌ Lỗi xử lý lệnh /pay: {e}", exc_info=True)
        await update.message.reply_text(
            "❌ Đã xảy ra lỗi khi tạo mã QR. Vui lòng thử lại sau."
        )


async def chart_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Xử lý lệnh /chart - Vẽ biểu đồ tròn chi tiêu"""
    logger.info(f"📨 Nhận lệnh /chart từ user: {update.effective_user.id}")
    
    try:
        # Đọc dữ liệu tháng này
        df = get_monthly_data()
        
        if df.empty:
            response = "❌ Tháng này chưa có dữ liệu chi tiêu.\n\nHãy thêm một vài giao dịch trước nhé!"
            await update.message.reply_text(response, parse_mode=ParseMode.MARKDOWN)
            return
        
        # Tính tổng theo phân loại
        category_totals = df.groupby('Phân loại')['Số tiền'].sum().sort_values(ascending=False)
        
        if category_totals.empty:
            response = "❌ Không có dữ liệu để vẽ biểu đồ."
            await update.message.reply_text(response, parse_mode=ParseMode.MARKDOWN)
            return
        
        # Vẽ biểu đồ tròn (Donut Chart)
        logger.info("🎨 Đang vẽ biểu đồ...")
        
        # Cấu hình style
        plt.style.use('default')
        sns.set_palette("pastel")
        
        fig, ax = plt.subplots(figsize=(10, 8))
        
        # Màu pastel đẹp mắt
        colors = ['#FFB6C1', '#87CEEB', '#98FB98', '#F0E68C', '#DDA0DD', '#FFA07A', '#20B2AA']
        
        # Vẽ donut chart
        wedges, texts, autotexts = ax.pie(
            category_totals.values,
            labels=category_totals.index,
            autopct='%1.1f%%',
            startangle=90,
            colors=colors[:len(category_totals)],
            pctdistance=0.85,
            textprops={'fontsize': 12, 'weight': 'bold'}
        )
        
        # Tạo hiệu ứng donut (khoảng trống ở giữa)
        centre_circle = plt.Circle((0, 0), 0.70, fc='white')
        ax.add_artist(centre_circle)
        
        # Thêm thông tin tổng ở giữa
        total_amount = category_totals.sum()
        ax.text(0, 0, f'Tổng:\n{total_amount:,}đ', 
                ha='center', va='center', 
                fontsize=16, weight='bold', color='#333333')
        
        # Tiêu đề
        now = datetime.now()
        ax.set_title(f'Chi Tiêu Tháng {now.month}/{now.year}', 
                    fontsize=18, weight='bold', pad=20)
        
        # Điều chỉnh layout
        plt.tight_layout()
        
        # Lưu vào BytesIO
        img_buffer = io.BytesIO()
        plt.savefig(img_buffer, format='png', dpi=150, bbox_inches='tight')
        img_buffer.seek(0)
        plt.close()  # Đóng figure để giải phóng bộ nhớ
        
        logger.info("✅ Đã tạo biểu đồ thành công")
        
        # Gửi ảnh qua Telegram
        await update.message.reply_photo(
            photo=img_buffer,
            caption=f"📊 **Biểu đồ chi tiêu tháng {now.month}/{now.year}**\n\n"
                   f"💰 Tổng: **{total_amount:,}đ**",
            parse_mode=ParseMode.MARKDOWN
        )
        logger.info("✅ Đã gửi biểu đồ cho user")
        
    except Exception as e:
        logger.error(f"❌ Lỗi khi tạo biểu đồ: {e}", exc_info=True)
        error_msg = "❌ Đã xảy ra lỗi khi tạo biểu đồ. Vui lòng thử lại sau."
        await update.message.reply_text(error_msg)


# ==================== XUẤT BÁO CÁO EXCEL ====================
async def export_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Xử lý lệnh /export - Xuất báo cáo Excel"""
    logger.info(f"📨 Nhận lệnh /export từ user: {update.effective_user.id}")
    
    try:
        # Đọc dữ liệu tháng này
        df = get_monthly_data()
        
        if df.empty:
            response = "❌ Tháng này chưa có dữ liệu chi tiêu.\n\nHãy thêm một vài giao dịch trước nhé!"
            await update.message.reply_text(response, parse_mode=ParseMode.MARKDOWN)
            return
        
        logger.info("📊 Đang tạo file Excel...")
        
        # Tạo Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Chi Tieu Thang {datetime.now().month}"
        
        # Header style
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Header
        headers = ['Full Time', 'Ngày', 'Tháng', 'Năm', 'Tên món', 'Phân loại', 'Số tiền']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Dữ liệu
        for row_num, row_data in enumerate(df.values, 2):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                if col_num == 7:  # Cột Số tiền
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.alignment = Alignment(horizontal="left")
        
        # Điều chỉnh độ rộng cột
        column_widths = [20, 8, 8, 8, 25, 15, 15]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width
        
        # Thêm dòng tổng
        total_row = len(df) + 3
        ws.cell(row=total_row, column=5, value="TỔNG CỘNG:").font = Font(bold=True)
        ws.cell(row=total_row, column=7, value=df['Số tiền'].sum())
        ws.cell(row=total_row, column=7).number_format = '#,##0'
        ws.cell(row=total_row, column=7).font = Font(bold=True)
        ws.cell(row=total_row, column=7).alignment = Alignment(horizontal="right")
        
        # Lưu vào BytesIO
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        logger.info("✅ Đã tạo file Excel thành công")
        
        # Tên file
        now = datetime.now()
        filename = f"BaoCaoChiTieu_{now.month}_{now.year}.xlsx"
        
        # Gửi file qua Telegram
        await update.message.reply_document(
            document=excel_buffer,
            filename=filename,
            caption=f"📊 **Báo cáo chi tiêu tháng {now.month}/{now.year}**\n\n"
                   f"📝 Tổng số giao dịch: {len(df)}\n"
                   f"💰 Tổng tiền: **{df['Số tiền'].sum():,}đ**",
            parse_mode=ParseMode.MARKDOWN
        )
        logger.info("✅ Đã gửi file Excel cho user")
        
    except Exception as e:
        logger.error(f"❌ Lỗi khi xuất Excel: {e}", exc_info=True)
        error_msg = "❌ Đã xảy ra lỗi khi xuất báo cáo. Vui lòng thử lại sau."
        await update.message.reply_text(error_msg)


async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Xử lý tin nhắn từ user - Multi-Line Parsing"""
    global groq_disabled  # Khai báo global ở đầu hàm
    
    user_text = update.message.text
    user_id = update.effective_user.id
    
    logger.info("=" * 60)
    logger.info("📨 NHẬN TIN NHẮN MỚI")
    logger.info("=" * 60)
    logger.info(f"👤 User ID: {user_id}")
    logger.info(f"💬 Tin nhắn: '{user_text}'")
    logger.info("-" * 60)
    
    try:
        # Hybrid AI + Fallback: Thử dùng Groq AI trước, nếu lỗi thì dùng Regex
        groq_result = None
        
        # Lấy financial context
        context_data = get_financial_context()
        
        # Ưu tiên Groq: Chỉ thử Groq nếu client khả dụng, chưa bị disable, và ưu tiên Groq
        if groq_client and not groq_disabled and GROQ_PRIORITY:
            try:
                logger.info("🤖 Đang thử parse bằng Groq AI...")
                groq_result = parse_with_groq(user_text, context_data, input_type='text')
                logger.info("✅ Đã sử dụng Groq AI thành công")
            except Exception as e:
                error_str = str(e).lower()
                # Log chi tiết hơn cho lỗi quota
                if 'quota' in error_str or 'rate limit' in error_str or '429' in error_str:
                    logger.warning("=" * 60)
                    logger.warning("⚠️ GROQ QUOTA HẾT - TỰ ĐỘNG CHUYỂN SANG REGEX")
                    logger.warning("💡 Bot vẫn sẽ thử Groq ở lần tiếp theo (quota có thể reset)")
                    logger.warning("💡 Kiểm tra quota: https://console.groq.com/usage")
                    logger.warning("=" * 60)
                else:
                    logger.warning(f"⚠️ Groq AI thất bại: {e}")
                logger.info("🔄 Chuyển sang Regex Fallback...")
                groq_result = None
        elif groq_disabled:
            # Groq đã bị disable (nếu có), bỏ qua luôn
            logger.info("ℹ️ Groq đã bị tắt tạm thời. Sử dụng Regex...")
            groq_result = None
        
        # Xử lý kết quả từ Groq
        if groq_result:
            if groq_result['type'] == 'chat':
                # Chat mode: Chỉ trả lời, không lưu Sheet
                await update.message.reply_text(groq_result['response'], parse_mode=ParseMode.MARKDOWN)
                logger.info("✅ Đã gửi phản hồi chat cho user")
                logger.info("=" * 60)
                logger.info("✅ XỬ LÝ TIN NHẮN THÀNH CÔNG!")
                logger.info("=" * 60)
                return
            elif groq_result['type'] == 'expense':
                # Expense mode: Lưu vào Sheet và trả lời
                expenses = groq_result['expenses']
                ai_message = groq_result.get('message', '')
                
                # Lưu vào Sheet
                saved_expenses = save_expenses_to_sheet(expenses)
                
                # Tính toán chi tiêu tuần
                weekly_data = calculate_weekly_spend()
                week_total = weekly_data['total']
                remaining = weekly_data['remaining']
                percentage = weekly_data['percentage']
                current_weekday = datetime.now().weekday()  # 0=Monday, 6=Sunday
                
                # Tạo phản hồi đẹp
                if len(saved_expenses) == 1:
                    expense = saved_expenses[0]
                    response = f"✅ **Đã lưu:**\n"
                    response += f"• {expense['item']}: {expense['amount']:,}đ ({expense['category']})"
                else:
                    response = f"✅ **Đã lưu {len(saved_expenses)} khoản chi:**\n"
                    total = 0
                    for expense in saved_expenses:
                        response += f"• {expense['item']}: {expense['amount']:,}đ ({expense['category']})\n"
                        total += expense['amount']
                    response += f"\n💰 **Tổng cộng: {total:,}đ**"
                
                # Thêm message từ AI nếu có
                if ai_message:
                    response += f"\n\n💬 {ai_message}"
                
                # Thêm thông tin ngân sách tuần
                response += f"\n\n📊 **Tuần này:** {week_total:,}đ / {WEEKLY_LIMIT:,}đ"
                
                if remaining < 0:
                    # Đã lố ngân sách
                    over_budget = abs(remaining)
                    response += f"\n⚠️ **BÁO ĐỘNG:** Bạn đã tiêu lố {over_budget:,}đ so với định mức tuần!"
                else:
                    response += f" (Còn dư: {remaining:,}đ)"
                
                # Cảnh báo thông minh: Nếu tiêu quá 80% và mới Thứ 3 hoặc Thứ 4
                if percentage >= 80 and current_weekday <= 3:  # Monday=0, Tuesday=1, Wednesday=2, Thursday=3
                    day_names = ['Thứ 2', 'Thứ 3', 'Thứ 4', 'Thứ 5', 'Thứ 6', 'Thứ 7', 'Chủ Nhật']
                    current_day_name = day_names[current_weekday]
                    response += f"\n\n⚠️ **Cảnh báo:** Tiêu chậm thôi, mới {current_day_name} đấy! ({percentage:.1f}% đã dùng)"
                
                # Kiểm tra từ khóa lãng phí và thêm cảnh báo
                for expense in saved_expenses:
                    wasteful_warning = get_wasteful_warning(expense['item'])
                    if wasteful_warning:
                        response += f"\n\n🚨 {wasteful_warning}"
                        break  # Chỉ thêm 1 cảnh báo cho mỗi lần lưu
                
                await update.message.reply_text(response, parse_mode=ParseMode.MARKDOWN)
                logger.info("✅ Đã gửi phản hồi expense cho user")
                logger.info("=" * 60)
                logger.info("✅ XỬ LÝ TIN NHẮN THÀNH CÔNG!")
                logger.info("=" * 60)
                return
        
        # Fallback về Regex nếu AI không khả dụng hoặc lỗi
        # Kiểm tra xem có phải yêu cầu tạo QR không (pattern matching)
        qr_keywords = ['mã qr', 'qr code', 'mã chuyển khoản', 'tạo qr', 'qr', 'chuyển khoản']
        text_lower = user_text.lower()
        
        if any(keyword in text_lower for keyword in qr_keywords):
            # Có từ khóa QR, thử parse số tiền và nội dung
            logger.info("🔄 Phát hiện yêu cầu tạo QR (Regex Fallback)...")
            
            # Parse số tiền
            amount = parse_amount_for_split(user_text)
            
            if amount > 0:
                # Tìm nội dung (text sau số tiền)
                import re
                # Tìm pattern số tiền và lấy text sau đó
                amount_pattern = r'(\d+(?:\.\d+)?)\s*(?:k|ng|nghìn|tr|triệu|đ|d)'
                match = re.search(amount_pattern, text_lower)
                
                content = ""
                if match:
                    # Lấy text sau số tiền
                    end_pos = match.end()
                    remaining_text = user_text[end_pos:].strip()
                    # Loại bỏ các từ khóa không cần thiết
                    remaining_text = re.sub(r'\b(tạo|cho|tôi|cái|mã|qr|code|chuyển|khoản|mệnh|giá|nội|dung|là)\b', '', remaining_text, flags=re.IGNORECASE).strip()
                    if remaining_text:
                        content = remaining_text
                
                logger.info(f"💳 Regex parse QR: {amount:,}đ - '{content}'")
                
                # Tạo URL VietQR
                qr_url = generate_vietqr_url(amount, content)
                
                if qr_url:
                    try:
                        img_response = requests.get(qr_url, timeout=10)
                        if img_response.status_code == 200:
                            image_buffer = io.BytesIO(img_response.content)
                            image_buffer.seek(0)
                            
                            # Tạo caption
                            caption = (
                                f"💳 **Quét mã này bank cho sếp Lộc nha!**\n"
                                f"💰 **Số tiền:** {amount:,}đ\n"
                                f"🏦 **VPBank - 0375646013**\n"
                                f"👤 **LE PHUOC LOC**"
                            )
                            if content:
                                caption += f"\n📝 **Nội dung:** {content}"
                            
                            # Gửi ảnh QR code
                            await update.message.reply_photo(
                                photo=image_buffer,
                                caption=caption,
                                parse_mode=ParseMode.MARKDOWN
                            )
                            logger.info("✅ Đã gửi mã QR VietQR (Regex Fallback)")
                            
                            # Lưu vào memory
                            add_to_memory(user_id, 'user', user_text)
                            add_to_memory(user_id, 'bot', f"Đã tạo mã QR {amount:,}đ")
                            
                            logger.info("=" * 60)
                            logger.info("✅ XỬ LÝ TIN NHẮN THÀNH CÔNG!")
                            logger.info("=" * 60)
                            return
                    except Exception as e:
                        logger.error(f"❌ Lỗi tải ảnh QR: {e}", exc_info=True)
        
        # Fallback về Regex cho chi tiêu
        logger.info("🔄 Sử dụng Regex Fallback cho chi tiêu...")
        expenses = parse_multiple_items(user_text)
        logger.info("✅ Đã sử dụng Regex Parsing (Fallback)")
        
        # Lưu vào Sheet
        saved_expenses = save_expenses_to_sheet(expenses)
        
        # Tính toán chi tiêu tuần
        weekly_data = calculate_weekly_spend()
        week_total = weekly_data['total']
        remaining = weekly_data['remaining']
        percentage = weekly_data['percentage']
        current_weekday = datetime.now().weekday()  # 0=Monday, 6=Sunday
        
        # Tạo phản hồi đẹp
        if len(saved_expenses) == 1:
            expense = saved_expenses[0]
            response = f"✅ **Đã lưu:**\n"
            response += f"• {expense['item']}: {expense['amount']:,}đ ({expense['category']})"
        else:
            response = f"✅ **Đã lưu {len(saved_expenses)} khoản chi:**\n"
            total = 0
            for expense in saved_expenses:
                response += f"• {expense['item']}: {expense['amount']:,}đ ({expense['category']})\n"
                total += expense['amount']
            response += f"\n💰 **Tổng cộng: {total:,}đ**"
        
        # Thêm thông tin ngân sách tuần
        response += f"\n\n📊 **Tuần này:** {week_total:,}đ / {WEEKLY_LIMIT:,}đ"
        
        if remaining < 0:
            # Đã lố ngân sách
            over_budget = abs(remaining)
            response += f"\n⚠️ **BÁO ĐỘNG:** Bạn đã tiêu lố {over_budget:,}đ so với định mức tuần!"
        else:
            response += f" (Còn dư: {remaining:,}đ)"
        
        # Cảnh báo thông minh: Nếu tiêu quá 80% và mới Thứ 3 hoặc Thứ 4
        if percentage >= 80 and current_weekday <= 3:  # Monday=0, Tuesday=1, Wednesday=2, Thursday=3
            day_names = ['Thứ 2', 'Thứ 3', 'Thứ 4', 'Thứ 5', 'Thứ 6', 'Thứ 7', 'Chủ Nhật']
            current_day_name = day_names[current_weekday]
            response += f"\n\n⚠️ **Cảnh báo:** Tiêu chậm thôi, mới {current_day_name} đấy! ({percentage:.1f}% đã dùng)"
        
        # Kiểm tra từ khóa lãng phí và thêm cảnh báo
        for expense in saved_expenses:
            wasteful_warning = get_wasteful_warning(expense['item'])
            if wasteful_warning:
                response += f"\n\n🚨 {wasteful_warning}"
                break  # Chỉ thêm 1 cảnh báo cho mỗi lần lưu
        
        await update.message.reply_text(response, parse_mode=ParseMode.MARKDOWN)
        logger.info("✅ Đã gửi phản hồi thành công")
        logger.info("=" * 60)
        logger.info("✅ XỬ LÝ TIN NHẮN THÀNH CÔNG!")
        logger.info("=" * 60)
        
    except ValueError as e:
        error_str = str(e)
        logger.warning("=" * 60)
        logger.warning("⚠️ XỬ LÝ TIN NHẮN THẤT BẠI")
        logger.warning(f"📝 Lỗi: {error_str}")
        
        user_text_lower = user_text.lower().strip()
        
        # Phát hiện tin nhắn chào hỏi/thường
        greetings = ['alo', 'hello', 'hi', 'xin chào', 'chào', 'chao', 'hey', 'hế lô', 'he lo']
        is_greeting = any(greeting in user_text_lower for greeting in greetings)
        
        if is_greeting:
            # Trả lời thân thiện cho tin nhắn chào hỏi
            error_msg = (
                "👋 **Xin chào!**\n\n"
                "Tôi là bot quản lý chi tiêu của bạn! 💰\n\n"
                "📝 **Để thêm chi tiêu, hãy nhập:**\n"
                "• `phở 50k`\n"
                "• `cơm 35k, trà đá 5k`\n"
                "• `xăng 200k`\n\n"
                "💡 **Các lệnh khác:**\n"
                "• `/help` - Xem hướng dẫn đầy đủ\n"
                "• `/report` - Xem báo cáo chi tiêu\n"
                "• `/chart` - Xem biểu đồ\n"
                "• `/remind 21:30` - Đặt báo thức nhắc nhở"
            )
        else:
            # Tin nhắn không phải chào hỏi nhưng không parse được
            error_msg = (
                "❌ Em không hiểu, vui lòng nhập kiểu:\n"
                "• `Món ăn + số tiền`\n"
                "• `cơm 35k, trà 5k`\n\n"
                "**Ví dụ:**\n"
                "• `phở 50k`\n"
                "• `xăng 200k`\n"
                "• `cơm 35k, trà đá 5k`\n\n"
                "💡 Gõ `/help` để xem hướng dẫn đầy đủ"
            )
        
        await update.message.reply_text(error_msg, parse_mode=ParseMode.MARKDOWN)
        
    except Exception as e:
        logger.error("=" * 60)
        logger.error("❌ XỬ LÝ TIN NHẮN THẤT BẠI (Exception)")
        logger.error(f"📝 Lỗi: {e}")
        logger.error(f"💡 Chi tiết:", exc_info=True)
        
        error_msg = "❌ Đã xảy ra lỗi. Vui lòng thử lại sau."
        await update.message.reply_text(error_msg)


# ==================== XỬ LÝ VOICE (SPEECH-TO-TEXT) ====================
async def handle_voice(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Xử lý tin nhắn thoại - Chuyển giọng nói thành text"""
    global groq_disabled
    
    logger.info("=" * 60)
    logger.info("🎤 NHẬN TIN NHẮN THOẠI")
    logger.info("=" * 60)
    logger.info(f"👤 User ID: {update.effective_user.id}")
    
    try:
        voice = update.message.voice
        if not voice:
            await update.message.reply_text("❌ Không tìm thấy file audio.")
            return
        
        logger.info(f"📊 Voice file: {voice.duration}s, {voice.file_size} bytes")
        
        # Tải file về bộ nhớ đệm
        file = await context.bot.get_file(voice.file_id)
        voice_buffer = io.BytesIO()
        await file.download_to_memory(voice_buffer)
        voice_buffer.seek(0)
        
        logger.info("🔄 Đang chuyển đổi giọng nói thành text...")
        
        # Gửi lên Groq Whisper API
        if groq_client and not groq_disabled:
            try:
                voice_buffer.seek(0)
                
                # Gọi Groq Audio Transcription API
                logger.info("📤 Đang gửi audio lên Groq Whisper...")
                transcription = groq_client.audio.transcriptions.create(
                    file=('voice.m4a', voice_buffer, 'audio/m4a'),
                    model='whisper-large-v3',
                    response_format='json',
                    language='vi'
                )
                
                # Lấy text từ response
                text = transcription.text
                logger.info(f"🎤 Voice Text: '{text}'")
                
                if not text or not text.strip():
                    await update.message.reply_text(
                        "⚠️ Không thể nhận diện giọng nói.\n"
                        "💡 Vui lòng thử lại hoặc gửi tin nhắn text."
                    )
                    return
                
                # Thông báo đã nghe được
                await update.message.reply_text(f"🎤 Đã nghe: {text}")
                logger.info("🔄 Chuyển sang xử lý text...")
                
                # Lấy financial context
                context_data = get_financial_context()
                
                # Lấy chat history của user
                user_id = update.effective_user.id
                chat_history = format_chat_history(user_id)
                if chat_history:
                    logger.info(f"📚 Đã lấy chat history: {len(chat_history)} ký tự")
                
                # Gọi parse_with_groq với text đã chuyển đổi
                groq_result = parse_with_groq(text, context_data, input_type='text', chat_history=chat_history)
                
                # Xử lý kết quả từ Groq (giống như handle_text)
                if groq_result['type'] == 'chat':
                    # Chat mode: Trả lời bằng VOICE
                    bot_response = groq_result['response']
                    
                    # Gửi text response trước (để user biết bot đã hiểu)
                    await update.message.reply_text(f"🎤 Đã nghe: {text}\n\n💬 {bot_response}", parse_mode=ParseMode.MARKDOWN)
                    
                    # Tạo và gửi voice reply
                    if TTS_AVAILABLE:
                        try:
                            logger.info("🔊 Đang tạo voice reply...")
                            
                            # Tạo file âm thanh từ text
                            tts = gTTS(text=bot_response, lang='vi', slow=False)
                            
                            # Lưu vào file tạm
                            with tempfile.NamedTemporaryFile(delete=False, suffix='.mp3') as tmp_file:
                                tts.save(tmp_file.name)
                                
                                # Chuyển đổi MP3 sang OGG (Telegram yêu cầu)
                                audio = AudioSegment.from_mp3(tmp_file.name)
                                
                                # Lưu OGG vào buffer
                                ogg_buffer = io.BytesIO()
                                audio.export(ogg_buffer, format='ogg')
                                ogg_buffer.seek(0)
                                
                                # Gửi voice message
                                await update.message.reply_voice(voice=ogg_buffer)
                                
                                # Xóa file tạm
                                os.unlink(tmp_file.name)
                                
                                logger.info("✅ Đã gửi voice reply cho user")
                        except Exception as e:
                            logger.warning(f"⚠️ Không thể tạo voice reply: {e}")
                            # Nếu lỗi, vẫn gửi text như bình thường
                    else:
                        logger.info("ℹ️ TTS không khả dụng, chỉ gửi text")
                    
                    # Lưu vào memory: Câu hỏi (từ voice) và câu trả lời
                    add_to_memory(user_id, 'user', text)
                    add_to_memory(user_id, 'bot', bot_response)
                    
                    logger.info("✅ Đã gửi phản hồi chat cho user")
                    return
                elif groq_result['type'] == 'expense':
                    # Expense mode: Lưu vào Sheet và trả lời
                    expenses = groq_result['expenses']
                    ai_message = groq_result.get('message', '')
                    
                    # Lưu vào Sheet
                    saved_expenses = save_expenses_to_sheet(expenses)
                    
                    # Tính toán chi tiêu tuần
                    weekly_data = calculate_weekly_spend()
                    week_total = weekly_data['total']
                    remaining = weekly_data['remaining']
                    percentage = weekly_data['percentage']
                    current_weekday = datetime.now().weekday()
                    
                    # Tạo phản hồi
                    if len(saved_expenses) == 1:
                        expense = saved_expenses[0]
                        response = f"✅ **Đã lưu từ voice:**\n"
                        response += f"• {expense['item']}: {expense['amount']:,}đ ({expense['category']})"
                    else:
                        response = f"✅ **Đã lưu {len(saved_expenses)} khoản chi từ voice:**\n"
                        total = 0
                        for expense in saved_expenses:
                            response += f"• {expense['item']}: {expense['amount']:,}đ ({expense['category']})\n"
                            total += expense['amount']
                        response += f"\n💰 **Tổng cộng: {total:,}đ**"
                    
                    if ai_message:
                        response += f"\n\n💬 {ai_message}"
                    
                    response += f"\n\n📊 **Tuần này:** {week_total:,}đ / {WEEKLY_LIMIT:,}đ"
                    if remaining < 0:
                        over_budget = abs(remaining)
                        response += f"\n⚠️ **BÁO ĐỘNG:** Bạn đã tiêu lố {over_budget:,}đ so với định mức tuần!"
                    else:
                        response += f" (Còn dư: {remaining:,}đ)"
                    
                    # Cảnh báo thông minh
                    if percentage >= 80 and current_weekday <= 3:
                        day_names = ['Thứ 2', 'Thứ 3', 'Thứ 4', 'Thứ 5', 'Thứ 6', 'Thứ 7', 'Chủ Nhật']
                        current_day_name = day_names[current_weekday]
                        response += f"\n\n⚠️ **Cảnh báo:** Tiêu chậm thôi, mới {current_day_name} đấy! ({percentage:.1f}% đã dùng)"
                    
                    # Kiểm tra từ khóa lãng phí
                    for expense in saved_expenses:
                        wasteful_warning = get_wasteful_warning(expense['item'])
                        if wasteful_warning:
                            response += f"\n\n🚨 {wasteful_warning}"
                            break
                    
                    # Gửi text response trước
                    await update.message.reply_text(response, parse_mode=ParseMode.MARKDOWN)
                    
                    # Tạo và gửi voice reply
                    if TTS_AVAILABLE:
                        try:
                            logger.info("🔊 Đang tạo voice reply cho expense...")
                            
                            # Tạo text ngắn gọn cho voice
                            voice_text = f"Đã lưu {len(saved_expenses)} khoản chi"
                            if len(saved_expenses) == 1:
                                voice_text = f"Đã lưu {saved_expenses[0]['item']} {saved_expenses[0]['amount']:,}đ"
                            
                            # Tạo file âm thanh
                            tts = gTTS(text=voice_text, lang='vi', slow=False)
                            
                            # Lưu vào file tạm
                            with tempfile.NamedTemporaryFile(delete=False, suffix='.mp3') as tmp_file:
                                tts.save(tmp_file.name)
                                
                                # Chuyển đổi MP3 sang OGG
                                audio = AudioSegment.from_mp3(tmp_file.name)
                                
                                # Lưu OGG vào buffer
                                ogg_buffer = io.BytesIO()
                                audio.export(ogg_buffer, format='ogg')
                                ogg_buffer.seek(0)
                                
                                # Gửi voice message
                                await update.message.reply_voice(voice=ogg_buffer)
                                
                                # Xóa file tạm
                                os.unlink(tmp_file.name)
                                
                                logger.info("✅ Đã gửi voice reply cho expense")
                        except Exception as e:
                            logger.warning(f"⚠️ Không thể tạo voice reply: {e}")
                    
                    # Lưu vào memory: Câu hỏi (từ voice) và câu trả lời
                    add_to_memory(user_id, 'user', text)
                    add_to_memory(user_id, 'bot', response)
                    
                    logger.info("✅ Đã xử lý voice thành công")
                    return
                
            except Exception as e:
                logger.error(f"❌ Lỗi Groq Whisper: {e}", exc_info=True)
                await update.message.reply_text(
                    "⚠️ Không thể chuyển đổi giọng nói.\n"
                    "💡 Vui lòng thử lại hoặc gửi tin nhắn text thay thế."
                )
                return
        else:
            await update.message.reply_text(
                "⚠️ Groq AI chưa được kích hoạt.\n"
                "💡 Vui lòng gửi tin nhắn text thay thế."
            )
            return
        
    except Exception as e:
        logger.error(f"❌ Lỗi xử lý voice: {e}", exc_info=True)
        await update.message.reply_text("❌ Đã xảy ra lỗi khi xử lý tin nhắn thoại.")


# ==================== XỬ LÝ PHOTO (VISION) ====================
async def handle_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Xử lý ảnh - Trích xuất chi tiêu từ hóa đơn/menu"""
    global groq_disabled
    
    logger.info("=" * 60)
    logger.info("📷 NHẬN ẢNH")
    logger.info("=" * 60)
    logger.info(f"👤 User ID: {update.effective_user.id}")
    
    try:
        photo = update.message.photo
        if not photo:
            await update.message.reply_text("❌ Không tìm thấy ảnh.")
            return
        
        # Lấy ảnh có độ phân giải cao nhất (cuối cùng trong list)
        photo_file = photo[-1]
        logger.info(f"📊 Photo: {photo_file.width}x{photo_file.height}, {photo_file.file_size} bytes")
        
        # Tải ảnh về bộ nhớ đệm
        file = await context.bot.get_file(photo_file.file_id)
        image_buffer = io.BytesIO()
        await file.download_to_memory(image_buffer)
        image_buffer.seek(0)
        
        # Mã hóa ảnh sang Base64
        image_base64 = base64.b64encode(image_buffer.read()).decode('utf-8')
        logger.info("🔄 Đang gửi ảnh lên Groq Vision...")
        
        # Lấy financial context
        context_data = get_financial_context()
        
        # Gửi lên Groq Vision
        if groq_client and not groq_disabled:
            try:
                groq_result = parse_with_groq(image_base64, context_data, input_type='image')
                
                # Kiểm tra nếu là lỗi Vision (đã được xử lý trong parse_with_groq)
                if groq_result.get('type') == 'chat' and 'bảo trì' in groq_result.get('response', ''):
                    await update.message.reply_text(groq_result['response'], parse_mode=ParseMode.MARKDOWN)
                    logger.info("⚠️ Vision Model bảo trì - Đã thông báo user")
                    return
                
                # Xử lý kết quả
                if groq_result['type'] == 'expense':
                    expenses = groq_result['expenses']
                    ai_message = groq_result.get('message', '')
                    
                    # Lưu vào Sheet
                    saved_expenses = save_expenses_to_sheet(expenses)
                    
                    # Tính toán chi tiêu tuần
                    weekly_data = calculate_weekly_spend()
                    week_total = weekly_data['total']
                    remaining = weekly_data['remaining']
                    
                    # Tạo phản hồi
                    if len(saved_expenses) == 1:
                        expense = saved_expenses[0]
                        response = f"✅ **Đã lưu từ ảnh:**\n"
                        response += f"• {expense['item']}: {expense['amount']:,}đ ({expense['category']})"
                    else:
                        response = f"✅ **Đã lưu {len(saved_expenses)} khoản chi từ ảnh:**\n"
                        total = 0
                        for expense in saved_expenses:
                            response += f"• {expense['item']}: {expense['amount']:,}đ ({expense['category']})\n"
                            total += expense['amount']
                        response += f"\n💰 **Tổng cộng: {total:,}đ**"
                    
                    if ai_message:
                        response += f"\n\n💬 {ai_message}"
                    
                    response += f"\n\n📊 **Tuần này:** {week_total:,}đ / {WEEKLY_LIMIT:,}đ"
                    if remaining < 0:
                        response += f"\n⚠️ **Đã vượt quá:** {abs(remaining):,}đ"
                    else:
                        response += f" (Còn dư: {remaining:,}đ)"
                    
                    await update.message.reply_text(response, parse_mode=ParseMode.MARKDOWN)
                    logger.info("✅ Đã xử lý ảnh thành công")
                    
                elif groq_result['type'] == 'chat':
                    await update.message.reply_text(groq_result['response'], parse_mode=ParseMode.MARKDOWN)
                    
            except Exception as e:
                error_str = str(e).lower()
                error_code = getattr(e, 'status_code', None)
                
                # Kiểm tra lỗi BadRequest (400) hoặc model decommissioned
                if (
                    '400' in str(e) or 
                    'bad request' in error_str or 
                    'decommissioned' in error_str or
                    'not found' in error_str or
                    error_code == 400
                ):
                    logger.warning("=" * 60)
                    logger.warning("⚠️ Vision Model đang bảo trì hoặc bị đổi tên")
                    logger.warning(f"⚠️ Error: {e}")
                    logger.warning("=" * 60)
                    await update.message.reply_text(
                        "❌ Tính năng đọc ảnh đang bảo trì do Groq thay đổi Model.\n"
                        "💡 Vui lòng nhập tay nhé sếp!"
                    )
                else:
                    logger.error(f"❌ Lỗi xử lý ảnh với Groq: {e}", exc_info=True)
                    await update.message.reply_text(
                        "❌ Không thể đọc thông tin từ ảnh.\n"
                        "💡 Vui lòng gửi lại ảnh rõ hơn hoặc nhập text thay thế."
                    )
        else:
            await update.message.reply_text(
                "⚠️ Groq Vision chưa được kích hoạt.\n"
                "💡 Vui lòng nhập text thay thế."
            )
        
    except Exception as e:
        logger.error(f"❌ Lỗi xử lý photo: {e}", exc_info=True)
        await update.message.reply_text("❌ Đã xảy ra lỗi khi xử lý ảnh.")


# ==================== HELPER FUNCTIONS FOR INTENT HANDLING ====================
async def send_alarm_spam(context: ContextTypes.DEFAULT_TYPE):
    """Hàm spam báo thức - Gửi tin nhắn lặp lại mỗi 10 giây"""
    chat_id = context.job.chat_id
    username = context.job.data.get('username', 'sếp')
    
    try:
        spam_message = f"Dậy đi! Dậy đi! 📢 @{username}"
        await context.bot.send_message(
            chat_id=chat_id,
            text=spam_message,
            parse_mode=ParseMode.MARKDOWN
        )
        logger.info(f"📢 Đã gửi spam báo thức cho chat {chat_id}")
    except Exception as e:
        logger.error(f"❌ Lỗi khi gửi spam báo thức: {e}")


async def trigger_alarm(context: ContextTypes.DEFAULT_TYPE):
    """Hàm kích hoạt báo thức - Gửi tin nhắn đầu tiên và bắt đầu spam"""
    chat_id = context.job.chat_id
    username = context.job.data.get('username', 'sếp')
    note = context.job.data.get('note', 'Dậy ngay sếp ơi')
    
    try:
        # Gửi tin nhắn báo thức đầu tiên
        alarm_message = f"🚨 **BÁO THỨC:** {note}! Dậy ngay sếp ơi @{username}"
        await context.bot.send_message(
            chat_id=chat_id,
            text=alarm_message,
            parse_mode=ParseMode.MARKDOWN
        )
        logger.info(f"🚨 Đã gửi báo thức cho chat {chat_id}")
        
        # Kích hoạt spam mode - Tạo job lặp lại mỗi 30 giây
        job_queue = context.application.job_queue
        if job_queue:
            # Lưu job spam vào chat_data để quản lý
            if 'alarm_spam_jobs' not in context.chat_data:
                context.chat_data['alarm_spam_jobs'] = []
            
            spam_job = job_queue.run_repeating(
                send_alarm_spam,
                interval=10,  # 30 giây
                first=0,  # Bắt đầu ngay
                chat_id=chat_id,
                data={'username': username}
            )
            
            context.chat_data['alarm_spam_jobs'].append(spam_job)
            logger.info(f"📢 Đã kích hoạt spam mode cho chat {chat_id}")
            
    except Exception as e:
        logger.error(f"❌ Lỗi khi kích hoạt báo thức: {e}", exc_info=True)


async def handle_alarm_intent(update: Update, context: ContextTypes.DEFAULT_TYPE, alarm_data: dict):
    """Xử lý ALARM intent - Đặt báo thức với spam mode"""
    user_id = str(update.effective_user.id)
    time_str = alarm_data.get('time', '')
    note = alarm_data.get('note', 'Dậy ngay sếp ơi')
    username = update.effective_user.username or update.effective_user.first_name or 'sếp'
    
    if not time_str:
        await update.message.reply_text(
            "❌ Không thể xác định thời gian. Vui lòng nhập: `/remind HH:MM`",
            parse_mode=ParseMode.MARKDOWN
        )
        return
    
    try:
        # Parse thời gian (HH:MM)
        time_parts = time_str.split(':')
        if len(time_parts) != 2:
            raise ValueError("Sai định dạng")
        
        hour = int(time_parts[0])
        minute = int(time_parts[1])
        
        if hour < 0 or hour > 23 or minute < 0 or minute > 59:
            raise ValueError("Giờ không hợp lệ")
        
        # Tính toán thời gian đến giờ hẹn
        now = datetime.now()
        alarm_time = dt_time(hour, minute)
        alarm_datetime = datetime.combine(now.date(), alarm_time)
        
        # Nếu giờ hẹn đã qua hôm nay, đặt cho ngày mai
        if alarm_datetime <= now:
            alarm_datetime += timedelta(days=1)
        
        # Tính số giây đến giờ hẹn
        seconds_until_alarm = (alarm_datetime - now).total_seconds()
        
        # Lưu reminder
        user_reminders[user_id] = {
            'hour': hour, 
            'minute': minute, 
            'note': note,
            'chat_id': update.effective_chat.id
        }
        save_reminders()
        
        # Lên lịch job báo thức (chạy 1 lần)
        job_queue = context.application.job_queue
        if job_queue:
            # Xóa job cũ nếu có
            current_jobs = job_queue.get_jobs_by_name(f"alarm_{user_id}")
            for job in current_jobs:
                job.schedule_removal()
            
            # Tạo job mới - chạy 1 lần vào giờ hẹn
            job_queue.run_once(
                trigger_alarm,
                when=seconds_until_alarm,
                name=f"alarm_{user_id}",
                chat_id=update.effective_chat.id,
                data={'username': username, 'note': note}
            )
        
        response = (
            f"✅ **Đã đặt báo thức thành công!**\n\n"
            f"🔔 Bot sẽ báo thức lúc **{hour:02d}:{minute:02d}**\n"
            f"📝 Nội dung: {note}\n"
            f"📢 **Spam mode:** Bot sẽ spam mỗi 30 giây cho đến khi bạn dừng\n\n"
            f"💡 Gõ 'Dậy rồi' hoặc 'Thôi đừng spam nữa' để tắt báo thức"
        )
        await update.message.reply_text(response, parse_mode=ParseMode.MARKDOWN)
        logger.info(f"✅ Đã đặt báo thức với spam mode: {hour:02d}:{minute:02d} (sau {seconds_until_alarm:.0f} giây)")
        
    except (ValueError, IndexError) as e:
        logger.error(f"❌ Lỗi parse thời gian: {e}", exc_info=True)
        await update.message.reply_text(
            "❌ Không thể đặt báo thức. Vui lòng nhập: `/remind HH:MM`",
            parse_mode=ParseMode.MARKDOWN
        )
    except Exception as e:
        logger.error(f"❌ Lỗi khi đặt báo thức: {e}", exc_info=True)
        await update.message.reply_text(
            "❌ Đã xảy ra lỗi khi đặt báo thức. Vui lòng thử lại sau.",
            parse_mode=ParseMode.MARKDOWN
        )


# ==================== XỬ LÝ TEXT (INTENT-BASED) ====================
async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Xử lý tin nhắn text - Sử dụng Intent Classification"""
    global groq_disabled
    
    user_text = update.message.text
    user_id = update.effective_user.id
    
    logger.info("=" * 60)
    logger.info("📨 NHẬN TIN NHẮN TEXT")
    logger.info("=" * 60)
    logger.info(f"👤 User ID: {user_id}")
    logger.info(f"💬 Tin nhắn: '{user_text}'")
    logger.info("-" * 60)
    
    try:
        # BƯỚC 1: Intent Classification với AI
        chat_history = format_chat_history(user_id)
        intent_result = None
        
        if groq_client and not groq_disabled:
            try:
                intent_result = classify_intent_with_ai(user_text, chat_history, groq_client)
                logger.info(f"🧠 Intent được phân loại: {intent_result['intent']}")
            except Exception as e:
                logger.warning(f"⚠️ Intent Classification thất bại: {e}")
                intent_result = None
        
        # Nếu không có intent, fallback về logic cũ
        if not intent_result:
            logger.info("🔄 Fallback về logic cũ (không có Intent Classification)")
            await handle_text_fallback(update, context)
            return
        
        intent = intent_result.get('intent', 'CHAT')
        intent_data = intent_result.get('data', {})
        
        # BƯỚC 2: Định tuyến dựa trên Intent (match/case pattern)
        try:
            if intent == 'EXPENSE':
                # Xử lý chi tiêu
                await handle_expense_intent(update, context, intent_data)
                
            elif intent == 'ALARM':
                # Xử lý đặt báo thức với spam mode
                await handle_alarm_intent(update, context, intent_data)
                
            elif intent == 'STOP':
                # Xử lý dừng báo thức spam
                await handle_stop_intent(update, context)
                
            elif intent == 'QR' or intent == 'QR_CODE':
                # Xử lý tạo QR code
                await handle_qr_intent(update, context, intent_data)
                
            elif intent == 'STOP':
                # Xử lý dừng báo thức spam
                await handle_stop_intent(update, context)
                
            elif intent == 'SEARCH':
                # Xử lý tìm kiếm Google
                await handle_search_intent(update, context, intent_data)
                
            elif intent == 'CHAT':
                # Xử lý chat thông thường
                await handle_chat_intent(update, context, intent_data, user_text, user_id, chat_history)
                
            else:
                # Fallback về chat
                logger.warning(f"⚠️ Intent không xác định: {intent}, chuyển về CHAT")
                await handle_chat_intent(update, context, intent_data, user_text, user_id, chat_history)
                
        except Exception as e:
            logger.error(f"❌ Lỗi khi xử lý intent {intent}: {e}", exc_info=True)
            # Fallback về chat với thông báo lỗi khéo léo
            error_response = (
                "Xin lỗi sếp, em gặp chút vấn đề kỹ thuật. "
                "Vui lòng thử lại hoặc mô tả rõ hơn yêu cầu của sếp nhé! 😊"
            )
            await update.message.reply_text(error_response, parse_mode=ParseMode.MARKDOWN)
            add_to_memory(user_id, 'user', user_text)
            add_to_memory(user_id, 'bot', error_response)
        
        logger.info("=" * 60)
        logger.info("✅ XỬ LÝ TIN NHẮN THÀNH CÔNG!")
        logger.info("=" * 60)
        
    except Exception as e:
        logger.error("=" * 60)
        logger.error("❌ XỬ LÝ TIN NHẮN THẤT BẠI (Exception)")
        logger.error(f"📝 Lỗi: {e}")
        logger.error(f"💡 Chi tiết:", exc_info=True)
        
        # Fallback về logic cũ
        await handle_text_fallback(update, context)


# ==================== INTENT HANDLERS ====================
async def handle_expense_intent(update: Update, context: ContextTypes.DEFAULT_TYPE, intent_data: dict):
    """Xử lý EXPENSE intent"""
    user_text = update.message.text
    user_id = update.effective_user.id
    
    logger.info("💰 Xử lý EXPENSE intent...")
    
    # Lấy thông tin từ intent_data
    amount = intent_data.get('amount', 0)
    item = intent_data.get('item', '')
    expense_date = intent_data.get('date')
    
    # Nếu không có đủ thông tin, fallback về logic cũ
    if not amount or not item:
        logger.warning("⚠️ Intent data không đủ, fallback về logic cũ")
        await handle_text_fallback(update, context)
        return
    
    # Tạo expense object
    expense = {
        'item': item,
        'amount': amount,
        'category': auto_categorize(item)
    }
    if expense_date:
        expense['date'] = expense_date
    
    # Lưu vào Sheet
    saved_expenses = save_expenses_to_sheet([expense])
    
    # Tính toán và trả lời
    weekly_data = calculate_weekly_spend()
    week_total = weekly_data['total']
    remaining = weekly_data['remaining']
    
    response = f"✅ **Đã lưu:**\n"
    response += f"• {expense['item']}: {expense['amount']:,}đ ({expense['category']})"
    response += f"\n\n📊 **Tuần này:** {week_total:,}đ / {WEEKLY_LIMIT:,}đ"
    
    if remaining < 0:
        over_budget = abs(remaining)
        response += f"\n⚠️ **BÁO ĐỘNG:** Bạn đã tiêu lố {over_budget:,}đ!"
    else:
        response += f" (Còn dư: {remaining:,}đ)"
    
    await update.message.reply_text(response, parse_mode=ParseMode.MARKDOWN)
    add_to_memory(user_id, 'user', user_text)
    add_to_memory(user_id, 'bot', response)


async def handle_stop_intent(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Xử lý STOP intent - Dừng báo thức spam"""
    user_id = str(update.effective_user.id)
    chat_id = update.effective_chat.id
    username = update.effective_user.username or update.effective_user.first_name or 'sếp'
    
    try:
        logger.info(f"🛑 Xử lý STOP intent cho user {user_id}")
        
        # Dừng các job spam trong chat_data
        spam_jobs_stopped = 0
        if 'alarm_spam_jobs' in context.chat_data:
            spam_jobs = context.chat_data.get('alarm_spam_jobs', [])
            for job in spam_jobs[:]:  # Copy list để tránh lỗi khi modify
                try:
                    job.schedule_removal()
                    spam_jobs_stopped += 1
                    logger.info(f"✅ Đã dừng spam job: {job.name}")
                except Exception as e:
                    logger.warning(f"⚠️ Không thể dừng job: {e}")
            
            # Xóa danh sách spam jobs
            context.chat_data['alarm_spam_jobs'] = []
        
        # Dừng các job báo thức chính
        job_queue = context.application.job_queue
        if job_queue:
            # Tìm và dừng job báo thức
            alarm_jobs = job_queue.get_jobs_by_name(f"alarm_{user_id}")
            for job in alarm_jobs:
                try:
                    job.schedule_removal()
                    logger.info(f"✅ Đã dừng alarm job: {job.name}")
                except Exception as e:
                    logger.warning(f"⚠️ Không thể dừng alarm job: {e}")
        
        # Xóa reminder nếu có
        if user_id in user_reminders:
            del user_reminders[user_id]
            save_reminders()
        
        response = "✅ Ok, đã tắt báo thức. Chúc sếp ngày mới năng lượng! ⚡"
        await update.message.reply_text(response, parse_mode=ParseMode.MARKDOWN)
        
        add_to_memory(update.effective_user.id, 'user', update.message.text)
        add_to_memory(update.effective_user.id, 'bot', response)
        
        logger.info(f"✅ Đã dừng {spam_jobs_stopped} spam job(s) cho user {user_id}")
        
    except Exception as e:
        logger.error(f"❌ Lỗi khi dừng báo thức: {e}", exc_info=True)
        await update.message.reply_text(
            "❌ Đã xảy ra lỗi khi dừng báo thức. Vui lòng thử lại sau.",
            parse_mode=ParseMode.MARKDOWN
        )


async def handle_qr_intent(update: Update, context: ContextTypes.DEFAULT_TYPE, intent_data: dict):
    """Xử lý QR_CODE intent"""
    user_id = update.effective_user.id
    amount = intent_data.get('amount', 0)
    content = intent_data.get('content', '')
    
    logger.info(f"💳 Xử lý QR_CODE intent: {amount:,}đ - '{content}'")
    
    if not amount or amount <= 0:
        await update.message.reply_text(
            "❌ Không thể xác định số tiền. Vui lòng nhập: `/pay [số tiền] [nội dung]`",
            parse_mode=ParseMode.MARKDOWN
        )
        return
    
    # Tạo QR code
    qr_url = generate_vietqr_url(amount, content)
    
    if not qr_url:
        await update.message.reply_text("❌ Không thể tạo mã QR. Vui lòng thử lại sau.")
        return
    
    # Tải và gửi ảnh QR
    try:
        import requests
        import io
        img_response = requests.get(qr_url, timeout=10)
        if img_response.status_code == 200:
            image_buffer = io.BytesIO(img_response.content)
            image_buffer.seek(0)
            
            caption = (
                f"💳 **Quét mã này bank cho sếp Lộc nha!**\n"
                f"💰 **Số tiền:** {amount:,}đ\n"
                f"🏦 **VPBank - 0375646013**\n"
                f"👤 **LE PHUOC LOC**"
            )
            if content:
                caption += f"\n📝 **Nội dung:** {content}"
            
            await update.message.reply_photo(
                photo=image_buffer,
                caption=caption,
                parse_mode=ParseMode.MARKDOWN
            )
            
            add_to_memory(user_id, 'user', update.message.text)
            add_to_memory(user_id, 'bot', f"Đã tạo mã QR {amount:,}đ")
        else:
            await update.message.reply_text(f"❌ Không thể tải ảnh QR (HTTP {img_response.status_code})")
    except Exception as e:
        logger.error(f"❌ Lỗi tải ảnh QR: {e}", exc_info=True)
        await update.message.reply_text("❌ Không thể tải ảnh QR. Vui lòng thử lại sau.")


async def handle_search_intent(update: Update, context: ContextTypes.DEFAULT_TYPE, intent_data: dict):
    """Xử lý SEARCH intent"""
    user_text = update.message.text
    user_id = update.effective_user.id
    query = intent_data.get('query', user_text)
    
    logger.info(f"🔍 Xử lý SEARCH intent: '{query}'")
    
    if not query:
        await update.message.reply_text("❌ Không thể xác định từ khóa tìm kiếm.")
        return
    
    # Gọi Google Search
    try:
        search_results = google_search(query, num_results=5)
        
        if not search_results or "⚠️" in search_results:
            await update.message.reply_text(
                f"❌ {search_results if search_results else 'Không thể tìm kiếm. Vui lòng thử lại sau.'}"
            )
            return
        
        # Gửi kết quả lên Groq để tổng hợp
        if groq_client and not groq_disabled:
            try:
                synthesis_prompt = (
                    f"Đây là kết quả tìm kiếm từ Google cho câu hỏi: '{user_text}'\n\n"
                    f"KẾT QUẢ TÌM KIẾM:\n{search_results}\n\n"
                    f"Hãy trả lời câu hỏi ban đầu của user dựa trên thông tin tìm kiếm này. "
                    f"Trả lời ngắn gọn, chuyên nghiệp nhưng thân thiện (kiểu thư ký riêng). "
                    f"Trả về JSON: {{\"type\": \"chat\", \"response\": \"Câu trả lời...\"}}"
                )
                
                final_result = parse_with_groq(synthesis_prompt, "", input_type='text', chat_history="")
                
                if final_result.get('type') == 'chat':
                    bot_response = final_result.get('response', search_results)
                    await update.message.reply_text(bot_response, parse_mode=ParseMode.MARKDOWN)
                    add_to_memory(user_id, 'user', user_text)
                    add_to_memory(user_id, 'bot', bot_response)
                    return
            except Exception as e:
                logger.warning(f"⚠️ Groq synthesis thất bại: {e}")
        
        # Fallback: Gửi kết quả trực tiếp
        await update.message.reply_text(
            f"🔍 **Kết quả tìm kiếm:**\n\n{search_results}",
            parse_mode=ParseMode.MARKDOWN
        )
        add_to_memory(user_id, 'user', user_text)
        add_to_memory(user_id, 'bot', search_results)
        
    except Exception as e:
        logger.error(f"❌ Lỗi Google Search: {e}", exc_info=True)
        await update.message.reply_text(
            "⚠️ Không thể tìm kiếm lúc này. Vui lòng thử lại sau.",
            parse_mode=ParseMode.MARKDOWN
        )


async def handle_chat_intent(update: Update, context: ContextTypes.DEFAULT_TYPE, intent_data: dict, 
                            user_text: str, user_id: int, chat_history: str):
    """Xử lý CHAT intent"""
    logger.info("💬 Xử lý CHAT intent...")
    
    # Lấy financial context
    context_data = get_financial_context()
    
    # Gọi AI để trả lời
    if groq_client and not groq_disabled:
        try:
            # Cập nhật system prompt để bot trả lời ngắn gọn, chuyên nghiệp nhưng thân thiện
            reply_instruction = intent_data.get('reply_instruction', '')
            
            groq_result = parse_with_groq(
                user_text, 
                context_data, 
                input_type='text', 
                chat_history=chat_history
            )
            
            if groq_result and groq_result.get('type') == 'chat':
                bot_response = groq_result.get('response', 'Xin lỗi, em không hiểu câu hỏi này.')
                await update.message.reply_text(bot_response, parse_mode=ParseMode.MARKDOWN)
                add_to_memory(user_id, 'user', user_text)
                add_to_memory(user_id, 'bot', bot_response)
                return
        except Exception as e:
            logger.warning(f"⚠️ Groq chat thất bại: {e}")
    
    # Fallback: Trả lời thân thiện
    friendly_response = (
        "👋 Xin chào! Em là bot quản lý chi tiêu của sếp Lộc.\n\n"
        "💡 **Em có thể giúp:**\n"
        "• Ghi chép chi tiêu (VD: `phở 50k`, `cơm 35k`)\n"
        "• Xem báo cáo tài chính (`/report`)\n"
        "• Tạo mã QR chuyển khoản (`/pay 50k nội dung`)\n"
        "• Trả lời câu hỏi về tài chính\n\n"
        "💬 **Hoặc gõ `/help` để xem hướng dẫn đầy đủ**"
    )
    await update.message.reply_text(friendly_response, parse_mode=ParseMode.MARKDOWN)
    add_to_memory(user_id, 'user', user_text)
    add_to_memory(user_id, 'bot', friendly_response)


# ==================== FALLBACK HANDLER (LOGIC CŨ) ====================
async def handle_text_fallback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Fallback về logic cũ nếu Intent Classification thất bại"""
    global groq_disabled
    
    user_text = update.message.text
    user_id = update.effective_user.id
    
    logger.info("🔄 Sử dụng Fallback Handler (Logic cũ)...")
    
    try:
        # BƯỚC 1: Kiểm tra xem tin nhắn có chứa số tiền hay không
        contains_amount = has_amount(user_text)
        logger.info(f"🔍 Kiểm tra số tiền: {'CÓ' if contains_amount else 'KHÔNG'}")
        
        # Lấy financial context
        context_data = get_financial_context()
        
        # Lấy chat history của user
        chat_history = format_chat_history(user_id)
        if chat_history:
            logger.info(f"📚 Đã lấy chat history: {len(chat_history)} ký tự")
        
        # Hybrid AI + Fallback: Thử dùng Groq AI trước, nếu lỗi thì dùng Regex
        groq_result = None
        
        # Ưu tiên Groq: Chỉ thử Groq nếu client khả dụng, chưa bị disable, và ưu tiên Groq
        if groq_client and not groq_disabled and GROQ_PRIORITY:
            try:
                logger.info("🤖 Đang thử parse bằng Groq AI...")
                groq_result = parse_with_groq(user_text, context_data, input_type='text', chat_history=chat_history)
                logger.info("✅ Đã sử dụng Groq AI thành công")
            except Exception as e:
                error_str = str(e).lower()
                # Log chi tiết hơn cho lỗi quota
                if 'quota' in error_str or 'rate limit' in error_str or '429' in error_str:
                    logger.warning("=" * 60)
                    logger.warning("⚠️ GROQ QUOTA HẾT - TỰ ĐỘNG CHUYỂN SANG REGEX")
                    logger.warning("💡 Bot vẫn sẽ thử Groq ở lần tiếp theo (quota có thể reset)")
                    logger.warning("💡 Kiểm tra quota: https://console.groq.com/usage")
                    logger.warning("=" * 60)
                else:
                    logger.warning(f"⚠️ Groq AI thất bại: {e}")
                logger.info("🔄 Chuyển sang Regex Fallback...")
                groq_result = None
        elif groq_disabled:
            # Groq đã bị disable (nếu có), bỏ qua luôn
            logger.info("ℹ️ Groq đã bị tắt tạm thời. Sử dụng Regex...")
            groq_result = None
        
        # Xử lý kết quả từ Groq
        if groq_result:
            if groq_result['type'] == 'qr_request':
                # QR Request mode: Tạo mã QR chuyển khoản
                amount = groq_result['amount']
                content = groq_result.get('content', '')
                
                logger.info(f"💳 Tạo mã QR: {amount:,}đ - '{content}'")
                
                # Tạo URL VietQR
                qr_url = generate_vietqr_url(amount, content)
                
                if not qr_url:
                    await update.message.reply_text(
                        "❌ Không thể tạo mã QR. Vui lòng thử lại sau."
                    )
                    return
                
                # Tải ảnh QR từ URL
                try:
                    img_response = requests.get(qr_url, timeout=10)
                    if img_response.status_code == 200:
                        image_buffer = io.BytesIO(img_response.content)
                        image_buffer.seek(0)
                        
                        # Tạo caption
                        caption = (
                            f"💳 **Quét mã này bank cho sếp Lộc nha!**\n"
                            f"💰 **Số tiền:** {amount:,}đ\n"
                            f"🏦 **VPBank - 0375646013**\n"
                            f"👤 **LE PHUOC LOC**"
                        )
                        if content:
                            caption += f"\n📝 **Nội dung:** {content}"
                        
                        # Gửi ảnh QR code
                        await update.message.reply_photo(
                            photo=image_buffer,
                            caption=caption,
                            parse_mode=ParseMode.MARKDOWN
                        )
                        logger.info("✅ Đã gửi mã QR VietQR cho user")
                        
                        # Lưu vào memory
                        add_to_memory(user_id, 'user', user_text)
                        add_to_memory(user_id, 'bot', f"Đã tạo mã QR {amount:,}đ")
                        
                        logger.info("=" * 60)
                        logger.info("✅ XỬ LÝ TIN NHẮN THÀNH CÔNG!")
                        logger.info("=" * 60)
                        return
                    else:
                        await update.message.reply_text(
                            f"❌ Không thể tải ảnh QR (HTTP {img_response.status_code})"
                        )
                        return
                except Exception as e:
                    logger.error(f"❌ Lỗi tải ảnh QR: {e}", exc_info=True)
                    await update.message.reply_text(
                        "❌ Không thể tải ảnh QR. Vui lòng thử lại sau."
                    )
                    return
                    
            elif groq_result['type'] == 'search':
                # Search mode: Tìm kiếm Google và trả lời
                search_query = groq_result.get('query', '')
                if not search_query:
                    # Nếu không có query, trả lời như chat
                    await update.message.reply_text(
                        "⚠️ Không thể xác định từ khóa tìm kiếm.\n"
                        "💡 Vui lòng thử lại với câu hỏi cụ thể hơn."
                    )
                    return
                
                logger.info(f"🔍 AI yêu cầu tìm kiếm: '{search_query}'")
                
                # Gọi Google Search
                try:
                    search_results = google_search(search_query, num_results=5)
                    
                    if not search_results or "⚠️" in search_results:
                        # Lỗi search hoặc không có kết quả
                        await update.message.reply_text(
                            f"❌ {search_results if search_results else 'Không thể tìm kiếm. Vui lòng thử lại sau.'}"
                        )
                        return
                    
                    # Gửi kết quả tìm kiếm lên Groq lần thứ 2 để tổng hợp
                    logger.info("🤖 Đang gửi kết quả tìm kiếm lên Groq để tổng hợp...")
                    
                    synthesis_prompt = (
                        f"Đây là kết quả tìm kiếm từ Google cho câu hỏi: '{user_text}'\n\n"
                        f"KẾT QUẢ TÌM KIẾM:\n{search_results}\n\n"
                        f"Hãy trả lời câu hỏi ban đầu của user dựa trên thông tin tìm kiếm này. "
                        f"Trả lời ngắn gọn, chính xác, có thể tham khảo các link trong kết quả. "
                        f"Trả về JSON: {{\"type\": \"chat\", \"response\": \"Câu trả lời dựa trên kết quả tìm kiếm...\"}}"
                    )
                    
                    final_result = parse_with_groq(synthesis_prompt, "", input_type='text', chat_history="")
                    
                    if final_result.get('type') == 'chat':
                        bot_response = final_result.get('response', 'Không thể tổng hợp kết quả.')
                        await update.message.reply_text(bot_response, parse_mode=ParseMode.MARKDOWN)
                        
                        # Lưu vào memory
                        add_to_memory(user_id, 'user', user_text)
                        add_to_memory(user_id, 'bot', bot_response)
                        
                        logger.info("✅ Đã gửi phản hồi search cho user")
                        logger.info("=" * 60)
                        logger.info("✅ XỬ LÝ TIN NHẮN THÀNH CÔNG!")
                        logger.info("=" * 60)
                        return
                    else:
                        # Fallback: Gửi kết quả search trực tiếp
                        await update.message.reply_text(
                            f"🔍 **Kết quả tìm kiếm:**\n\n{search_results}",
                            parse_mode=ParseMode.MARKDOWN
                        )
                        add_to_memory(user_id, 'user', user_text)
                        add_to_memory(user_id, 'bot', search_results)
                        return
                        
                except Exception as e:
                    logger.error(f"❌ Lỗi Google Search: {e}", exc_info=True)
                    # Kiểm tra xem có phải do thiếu API keys không
                    if not GOOGLE_SEARCH_API_KEY or not GOOGLE_CSE_ID:
                        error_msg = (
                            "⚠️ **Tính năng tìm kiếm Google chưa được cấu hình.**\n\n"
                            "💡 Để sử dụng tính năng này, vui lòng:\n"
                            "1. Tạo Google Custom Search Engine\n"
                            "2. Lấy API Key từ Google Cloud Console\n"
                            "3. Thêm vào biến môi trường"
                        )
                    else:
                        error_msg = (
                            "⚠️ Không thể tìm kiếm lúc này.\n"
                            "💡 Vui lòng thử lại sau hoặc kiểm tra cấu hình Google Search API."
                        )
                    await update.message.reply_text(error_msg, parse_mode=ParseMode.MARKDOWN)
                    return
                    
            elif groq_result['type'] == 'chat':
                # Chat mode: Chỉ trả lời, không lưu Sheet
                bot_response = groq_result['response']
                await update.message.reply_text(bot_response, parse_mode=ParseMode.MARKDOWN)
                
                # Lưu vào memory: Câu hỏi và câu trả lời
                add_to_memory(user_id, 'user', user_text)
                add_to_memory(user_id, 'bot', bot_response)
                
                logger.info("✅ Đã gửi phản hồi chat cho user")
                logger.info("=" * 60)
                logger.info("✅ XỬ LÝ TIN NHẮN THÀNH CÔNG!")
                logger.info("=" * 60)
                return
            elif groq_result['type'] == 'expense':
                # Expense mode: Lưu vào Sheet và trả lời
                expenses = groq_result['expenses']
                ai_message = groq_result.get('message', '')
                
                # Lưu vào Sheet
                saved_expenses = save_expenses_to_sheet(expenses)
                
                # Tính toán chi tiêu tuần
                weekly_data = calculate_weekly_spend()
                week_total = weekly_data['total']
                remaining = weekly_data['remaining']
                percentage = weekly_data['percentage']
                current_weekday = datetime.now().weekday()  # 0=Monday, 6=Sunday
                
                # Tạo phản hồi đẹp
                if len(saved_expenses) == 1:
                    expense = saved_expenses[0]
                    response = f"✅ **Đã lưu:**\n"
                    response += f"• {expense['item']}: {expense['amount']:,}đ ({expense['category']})"
                else:
                    response = f"✅ **Đã lưu {len(saved_expenses)} khoản chi:**\n"
                    total = 0
                    for expense in saved_expenses:
                        response += f"• {expense['item']}: {expense['amount']:,}đ ({expense['category']})\n"
                        total += expense['amount']
                    response += f"\n💰 **Tổng cộng: {total:,}đ**"
                
                # Thêm message từ AI nếu có
                if ai_message:
                    response += f"\n\n💬 {ai_message}"
                
                # Thêm thông tin ngân sách tuần
                response += f"\n\n📊 **Tuần này:** {week_total:,}đ / {WEEKLY_LIMIT:,}đ"
                
                if remaining < 0:
                    # Đã lố ngân sách
                    over_budget = abs(remaining)
                    response += f"\n⚠️ **BÁO ĐỘNG:** Bạn đã tiêu lố {over_budget:,}đ so với định mức tuần!"
                else:
                    response += f" (Còn dư: {remaining:,}đ)"
                
                # Cảnh báo thông minh: Nếu tiêu quá 80% và mới Thứ 3 hoặc Thứ 4
                if percentage >= 80 and current_weekday <= 3:  # Monday=0, Tuesday=1, Wednesday=2, Thursday=3
                    day_names = ['Thứ 2', 'Thứ 3', 'Thứ 4', 'Thứ 5', 'Thứ 6', 'Thứ 7', 'Chủ Nhật']
                    current_day_name = day_names[current_weekday]
                    response += f"\n\n⚠️ **Cảnh báo:** Tiêu chậm thôi, mới {current_day_name} đấy! ({percentage:.1f}% đã dùng)"
                
                # Kiểm tra từ khóa lãng phí và thêm cảnh báo
                for expense in saved_expenses:
                    wasteful_warning = get_wasteful_warning(expense['item'])
                    if wasteful_warning:
                        response += f"\n\n🚨 {wasteful_warning}"
                        break  # Chỉ thêm 1 cảnh báo cho mỗi lần lưu
                
                await update.message.reply_text(response, parse_mode=ParseMode.MARKDOWN)
                logger.info("✅ Đã gửi phản hồi expense cho user")
                logger.info("=" * 60)
                logger.info("✅ XỬ LÝ TIN NHẮN THÀNH CÔNG!")
                logger.info("=" * 60)
                return
        
        # Fallback về Regex nếu AI không khả dụng hoặc lỗi
        # Kiểm tra xem có phải yêu cầu tạo QR không (pattern matching)
        qr_keywords = ['mã qr', 'qr code', 'mã chuyển khoản', 'tạo qr', 'qr', 'chuyển khoản']
        text_lower = user_text.lower()
        
        if any(keyword in text_lower for keyword in qr_keywords):
            # Có từ khóa QR, thử parse số tiền và nội dung
            logger.info("🔄 Phát hiện yêu cầu tạo QR (Regex Fallback)...")
            
            # Parse số tiền
            amount = parse_amount_for_split(user_text)
            
            if amount > 0:
                # Tìm nội dung (text sau số tiền)
                import re
                # Tìm pattern số tiền và lấy text sau đó
                amount_pattern = r'(\d+(?:\.\d+)?)\s*(?:k|ng|nghìn|tr|triệu|đ|d)'
                match = re.search(amount_pattern, text_lower)
                
                content = ""
                if match:
                    # Lấy text sau số tiền
                    end_pos = match.end()
                    remaining_text = user_text[end_pos:].strip()
                    # Loại bỏ các từ khóa không cần thiết
                    remaining_text = re.sub(r'\b(tạo|cho|tôi|cái|mã|qr|code|chuyển|khoản|mệnh|giá|nội|dung|là)\b', '', remaining_text, flags=re.IGNORECASE).strip()
                    if remaining_text:
                        content = remaining_text
                
                logger.info(f"💳 Regex parse QR: {amount:,}đ - '{content}'")
                
                # Tạo URL VietQR
                qr_url = generate_vietqr_url(amount, content)
                
                if qr_url:
                    try:
                        img_response = requests.get(qr_url, timeout=10)
                        if img_response.status_code == 200:
                            image_buffer = io.BytesIO(img_response.content)
                            image_buffer.seek(0)
                            
                            # Tạo caption
                            caption = (
                                f"💳 **Quét mã này bank cho sếp Lộc nha!**\n"
                                f"💰 **Số tiền:** {amount:,}đ\n"
                                f"🏦 **VPBank - 0375646013**\n"
                                f"👤 **LE PHUOC LOC**"
                            )
                            if content:
                                caption += f"\n📝 **Nội dung:** {content}"
                            
                            # Gửi ảnh QR code
                            await update.message.reply_photo(
                                photo=image_buffer,
                                caption=caption,
                                parse_mode=ParseMode.MARKDOWN
                            )
                            logger.info("✅ Đã gửi mã QR VietQR (Regex Fallback)")
                            
                            # Lưu vào memory
                            add_to_memory(user_id, 'user', user_text)
                            add_to_memory(user_id, 'bot', f"Đã tạo mã QR {amount:,}đ")
                            
                            logger.info("=" * 60)
                            logger.info("✅ XỬ LÝ TIN NHẮN THÀNH CÔNG!")
                            logger.info("=" * 60)
                            return
                    except Exception as e:
                        logger.error(f"❌ Lỗi tải ảnh QR: {e}", exc_info=True)
        
        # Fallback: Xử lý dựa trên việc có số tiền hay không
        if not contains_amount:
            # KHÔNG có số tiền → Coi là câu hỏi thông thường, gọi AI chat
            logger.info("💬 Tin nhắn không có số tiền → Xử lý như chat thông thường")
            
            # Thử gọi Groq để chat (nếu chưa gọi hoặc lỗi)
            if groq_client and not groq_disabled:
                try:
                    logger.info("🤖 Đang gọi Groq AI để trả lời câu hỏi...")
                    groq_result = parse_with_groq(user_text, context_data, input_type='text', chat_history=chat_history)
                    
                    if groq_result and groq_result.get('type') == 'chat':
                        bot_response = groq_result.get('response', 'Xin lỗi, em không hiểu câu hỏi này.')
                        await update.message.reply_text(bot_response, parse_mode=ParseMode.MARKDOWN)
                        
                        # Lưu vào memory
                        add_to_memory(user_id, 'user', user_text)
                        add_to_memory(user_id, 'bot', bot_response)
                        
                        logger.info("✅ Đã gửi phản hồi chat cho user")
                        logger.info("=" * 60)
                        logger.info("✅ XỬ LÝ TIN NHẮN THÀNH CÔNG!")
                        logger.info("=" * 60)
                        return
                except Exception as e:
                    logger.warning(f"⚠️ Groq AI chat thất bại: {e}")
            
            # Nếu Groq không khả dụng, trả lời thân thiện
            friendly_response = (
                "👋 Xin chào! Em là bot quản lý chi tiêu của sếp Lộc.\n\n"
                "💡 **Em có thể giúp:**\n"
                "• Ghi chép chi tiêu (VD: `phở 50k`, `cơm 35k`)\n"
                "• Xem báo cáo tài chính (`/report`)\n"
                "• Tạo mã QR chuyển khoản (`/pay 50k nội dung`)\n"
                "• Trả lời câu hỏi về tài chính\n\n"
                "📝 **Để thêm chi tiêu, hãy nhập:**\n"
                "• `phở 50k`\n"
                "• `cơm 35k, trà đá 5k`\n\n"
                "💬 **Hoặc gõ `/help` để xem hướng dẫn đầy đủ**"
            )
            await update.message.reply_text(friendly_response, parse_mode=ParseMode.MARKDOWN)
            
            # Lưu vào memory
            add_to_memory(user_id, 'user', user_text)
            add_to_memory(user_id, 'bot', friendly_response)
            
            logger.info("✅ Đã gửi phản hồi thân thiện cho user")
            logger.info("=" * 60)
            logger.info("✅ XỬ LÝ TIN NHẮN THÀNH CÔNG!")
            logger.info("=" * 60)
            return
        
        # CÓ số tiền → Xử lý như expense (Regex Fallback)
        logger.info("🔄 Sử dụng Regex Fallback cho chi tiêu...")
        expenses = parse_multiple_items(user_text)
        logger.info("✅ Đã sử dụng Regex Parsing (Fallback)")
        
        # Lưu vào Sheet
        saved_expenses = save_expenses_to_sheet(expenses)
        
        # Tính toán chi tiêu tuần
        weekly_data = calculate_weekly_spend()
        week_total = weekly_data['total']
        remaining = weekly_data['remaining']
        percentage = weekly_data['percentage']
        current_weekday = datetime.now().weekday()  # 0=Monday, 6=Sunday
        
        # Tạo phản hồi đẹp
        if len(saved_expenses) == 1:
            expense = saved_expenses[0]
            response = f"✅ **Đã lưu:**\n"
            response += f"• {expense['item']}: {expense['amount']:,}đ ({expense['category']})"
        else:
            response = f"✅ **Đã lưu {len(saved_expenses)} khoản chi:**\n"
            total = 0
            for expense in saved_expenses:
                response += f"• {expense['item']}: {expense['amount']:,}đ ({expense['category']})\n"
                total += expense['amount']
            response += f"\n💰 **Tổng cộng: {total:,}đ**"
        
        # Thêm thông tin ngân sách tuần
        response += f"\n\n📊 **Tuần này:** {week_total:,}đ / {WEEKLY_LIMIT:,}đ"
        
        if remaining < 0:
            # Đã lố ngân sách
            over_budget = abs(remaining)
            response += f"\n⚠️ **BÁO ĐỘNG:** Bạn đã tiêu lố {over_budget:,}đ so với định mức tuần!"
        else:
            response += f" (Còn dư: {remaining:,}đ)"
        
        # Cảnh báo thông minh: Nếu tiêu quá 80% và mới Thứ 3 hoặc Thứ 4
        if percentage >= 80 and current_weekday <= 3:  # Monday=0, Tuesday=1, Wednesday=2, Thursday=3
            day_names = ['Thứ 2', 'Thứ 3', 'Thứ 4', 'Thứ 5', 'Thứ 6', 'Thứ 7', 'Chủ Nhật']
            current_day_name = day_names[current_weekday]
            response += f"\n\n⚠️ **Cảnh báo:** Tiêu chậm thôi, mới {current_day_name} đấy! ({percentage:.1f}% đã dùng)"
        
        # Kiểm tra từ khóa lãng phí và thêm cảnh báo
        for expense in saved_expenses:
            wasteful_warning = get_wasteful_warning(expense['item'])
            if wasteful_warning:
                response += f"\n\n🚨 {wasteful_warning}"
                break  # Chỉ thêm 1 cảnh báo cho mỗi lần lưu
        
        await update.message.reply_text(response, parse_mode=ParseMode.MARKDOWN)
        
    except ValueError as e:
        error_str = str(e)
        logger.warning("=" * 60)
        logger.warning("⚠️ XỬ LÝ TIN NHẮN THẤT BẠI")
        logger.warning(f"📝 Lỗi: {error_str}")
        
        user_text_lower = user_text.lower().strip()
        
        # Phát hiện tin nhắn chào hỏi/thường
        greetings = ['alo', 'hello', 'hi', 'xin chào', 'chào', 'chao', 'hey', 'hế lô', 'he lo']
        is_greeting = any(greeting in user_text_lower for greeting in greetings)
        
        if is_greeting:
            # Trả lời thân thiện cho tin nhắn chào hỏi
            error_msg = (
                "👋 **Xin chào!**\n\n"
                "Tôi là bot quản lý chi tiêu của bạn! 💰\n\n"
                "📝 **Để thêm chi tiêu, hãy nhập:**\n"
                "• `phở 50k`\n"
                "• `cơm 35k, trà đá 5k`\n"
                "• `xăng 200k`\n\n"
                "💡 **Các lệnh khác:**\n"
                "• `/help` - Xem hướng dẫn đầy đủ\n"
                "• `/report` - Xem báo cáo chi tiêu\n"
                "• `/chart` - Xem biểu đồ\n"
                "• `/remind 21:30` - Đặt báo thức nhắc nhở"
            )
        else:
            # Tin nhắn không phải chào hỏi nhưng không parse được
            error_msg = (
                "❌ Em không hiểu, vui lòng nhập kiểu:\n"
                "• `Món ăn + số tiền`\n"
                "• `cơm 35k, trà 5k`\n\n"
                "**Ví dụ:**\n"
                "• `phở 50k`\n"
                "• `xăng 200k`\n"
                "• `cơm 35k, trà đá 5k`\n\n"
                "💡 Gõ `/help` để xem hướng dẫn đầy đủ"
            )
        
        await update.message.reply_text(error_msg, parse_mode=ParseMode.MARKDOWN)
        
    except Exception as e:
        logger.error("=" * 60)
        logger.error("❌ XỬ LÝ TIN NHẮN THẤT BẠI (Exception)")
        logger.error(f"📝 Lỗi: {e}")
        logger.error(f"💡 Chi tiết:", exc_info=True)
        
        error_msg = "❌ Đã xảy ra lỗi. Vui lòng thử lại sau."
        await update.message.reply_text(error_msg)


# ==================== HANDLE MESSAGE (WRAPPER) ====================
async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Wrapper để xử lý text messages"""
    await handle_text(update, context)


# ==================== HÀM CHÍNH ====================
def main():
    """Hàm chính để khởi chạy bot"""
    # Khởi động Keep Alive server cho Render.com
    keep_alive()
    logger.info("✅ Đã khởi động Keep Alive server (Flask)")
    
    logger.info("=" * 60)
    logger.info("🚀 KHỞI ĐỘNG BOT")
    logger.info("=" * 60)
    
    if worksheet is None:
        logger.critical("❌ CRITICAL ERROR: Không thể khởi động bot!")
        return
    
    application = Application.builder().token(TELEGRAM_TOKEN).build()
    logger.info("✅ Đã tạo Telegram Application")
    
    # Đăng ký handlers
    application.add_handler(CommandHandler("start", start_command))
    application.add_handler(CommandHandler("help", help_command))
    application.add_handler(CommandHandler("huongdan", help_command))  # Alias tiếng Việt
    application.add_handler(CommandHandler("report", report_command))
    application.add_handler(CommandHandler("thongke", report_command))
    application.add_handler(CommandHandler("chart", chart_command))
    application.add_handler(CommandHandler("export", export_command))
    application.add_handler(CommandHandler("undo", undo_command))
    application.add_handler(CommandHandler("delete", delete_command))
    application.add_handler(CommandHandler("xoa", delete_command))  # Alias tiếng Việt
    application.add_handler(CallbackQueryHandler(delete_callback, pattern="^delete_"))
    application.add_handler(CommandHandler("remind", remind_command))
    application.add_handler(CommandHandler("stopremind", stopremind_command))
    application.add_handler(CommandHandler("chia", chia_command))
    application.add_handler(CommandHandler("pay", pay_command))
    application.add_handler(CommandHandler("qr", pay_command))  # Alias cho /pay
    
    # Đăng ký handlers cho đa modal
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))
    application.add_handler(MessageHandler(filters.VOICE, handle_voice))
    application.add_handler(MessageHandler(filters.PHOTO, handle_photo))
    
    # Khôi phục reminders và lên lịch jobs
    job_queue = application.job_queue
    if job_queue:
        logger.info("🔔 Đang khôi phục reminders...")
        restored_count = 0
        for user_id, reminder_data in user_reminders.items():
            try:
                hour = reminder_data['hour']
                minute = reminder_data['minute']
                chat_id = reminder_data.get('chat_id')
                
                if chat_id:
                    reminder_time = dt_time(hour, minute)
                    job_queue.run_daily(
                        send_daily_reminder,
                        time=reminder_time,
                        name=f"reminder_{user_id}",
                        chat_id=chat_id
                    )
                    restored_count += 1
                    logger.info(f"  ✅ Đã khôi phục reminder cho user {user_id} lúc {hour:02d}:{minute:02d}")
                else:
                    logger.warning(f"  ⚠️ Reminder cho user {user_id} thiếu chat_id - cần đặt lại")
            except Exception as e:
                logger.warning(f"  ⚠️ Không thể khôi phục reminder cho user {user_id}: {e}")
        
        if restored_count > 0:
            logger.info(f"✅ Đã khôi phục {restored_count} reminders")
    logger.info("✅ Đã đăng ký handlers")
    
    logger.info("=" * 60)
    logger.info("✅ BOT ĐÃ SẴN SÀNG!")
    logger.info("=" * 60)
    logger.info(f"📊 Đã kết nối với Google Sheet")
    logger.info("🤖 Bot đang chạy và sẵn sàng nhận tin nhắn...")
    logger.info("💡 Enterprise Edition - Multi-Line, Charts, Excel Export Enabled")
    logger.info("=" * 60)
    
    # Xử lý lỗi với auto-recovery logic
    import time
    max_retries = 3
    retry_delay = 5  # seconds
    consecutive_failures = 0
    max_consecutive_failures = 10  # Sau 10 lần fail liên tiếp thì dừng
    
    while True:  # Infinite loop để bot luôn tự động recover
        try:
            # Trước khi start polling, thử dừng các webhook cũ (nếu có)
            try:
                bot_instance = application.bot
                bot_instance.delete_webhook(drop_pending_updates=True)
                logger.info("🔄 Đã xóa webhook cũ (nếu có)")
            except Exception as webhook_error:
                logger.debug(f"Không có webhook cũ để xóa: {webhook_error}")
            
            # Reset counter nếu thành công
            consecutive_failures = 0
            retry_delay = 5  # Reset delay
            
            # Start polling
            logger.info("🔄 Đang khởi động polling...")
            logger.info("💡 Bot sẽ tự động restart nếu gặp lỗi tạm thời")
            application.run_polling(
                allowed_updates=Update.ALL_TYPES, 
                drop_pending_updates=True,
                close_loop=False
            )
            # Nếu polling dừng (không có lỗi), restart lại
            logger.warning("⚠️ Polling đã dừng, đang restart...")
            time.sleep(2)
            
        except KeyboardInterrupt:
            logger.info("🛑 Bot đã được dừng bởi user (Ctrl+C)")
            break
        except Exception as e:
            consecutive_failures += 1
            error_str = str(e)
            error_type = type(e).__name__
            
            # Kiểm tra nếu là lỗi có thể retry
            retryable_errors = [
                "Conflict", "getUpdates", "NetworkError", "TimedOut", 
                "ConnectionError", "RetryAfter", "TelegramError"
            ]
            is_retryable = any(keyword in error_str or keyword in error_type for keyword in retryable_errors)
            
            if is_retryable:
                if consecutive_failures < max_consecutive_failures:
                    logger.warning("=" * 60)
                    logger.warning(f"⚠️ Lỗi phát hiện: {error_type}")
                    logger.warning(f"📝 Chi tiết: {error_str[:200]}")
                    logger.warning(f"🔄 Tự động restart sau {retry_delay} giây... (Lần thử {consecutive_failures}/{max_consecutive_failures})")
                    logger.warning("=" * 60)
                    time.sleep(retry_delay)
                    retry_delay = min(retry_delay * 1.5, 60)  # Exponential backoff, max 60s
                else:
                    logger.critical("=" * 60)
                    logger.critical(f"❌ CRITICAL: Đã fail {max_consecutive_failures} lần liên tiếp!")
                    logger.critical(f"📝 Lỗi cuối: {error_type}: {error_str[:200]}")
                    logger.critical("💡 Bot sẽ dừng. Vui lòng kiểm tra logs và restart thủ công.")
                    logger.critical("=" * 60)
                    break
            else:
                # Lỗi không thể retry (như API key sai, syntax error, etc.)
                logger.critical("=" * 60)
                logger.critical(f"❌ LỖI NGHIÊM TRỌNG: {error_type}")
                logger.critical(f"📝 {error_str[:300]}")
                logger.critical("💡 Đây là lỗi không thể tự động fix. Vui lòng kiểm tra code/config.")
                logger.critical("=" * 60)
                # Vẫn thử lại sau một khoảng thời gian dài (có thể là lỗi tạm thời của server)
                if consecutive_failures < max_consecutive_failures:
                    logger.warning(f"⏳ Đợi 30 giây rồi thử lại...")
                    time.sleep(30)
                else:
                    break


if __name__ == '__main__':
    main()
